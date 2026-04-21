"""Extract detailed URP_PARAM_CONFIG patterns from inspection SQL templates"""
import re, openpyxl, sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('F:/09_opencode/docs/my_poor_solution.xlsx', read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

# Collect all param SQL templates and analyze their patterns
all_param_sqls = []
param_value_pattern = re.compile(r"param_value\s*=\s*'([^']*)'")
param_code_pattern = re.compile(r"param_code\s*=\s*'([^']+)'")

for row in ws.iter_rows(min_row=2, values_only=True):
    if row and row[22]:
        sql = str(row[22])
        if 'URP_PARAM_CONFIG' in sql or '#[' in sql:
            alias = str(row[1]) if row[1] else ''
            dm_table = str(row[12]) if row[12] else ''
            dm_field = str(row[13]) if row[13] else ''
            check_code = str(row[3]) if row[3] else ''
            
            # Extract param_code references
            param_codes = list(set(param_code_pattern.findall(sql)))
            # Extract param_value references
            param_values = list(set(param_value_pattern.findall(sql)))
            # Check for NOT EXISTS pattern
            has_not_exists = 'not exists' in sql.lower() and 'URP_PARAM_CONFIG' in sql
            
            all_param_sqls.append({
                'alias': alias,
                'dm_table': dm_table,
                'dm_field': dm_field,
                'check_code': check_code,
                'param_codes': param_codes,
                'param_values': param_values,
                'has_not_exists': has_not_exists,
                'sql_len': len(sql),
                'sql': sql
            })

wb.close()

print(f"Total param SQL templates: {len(all_param_sqls)}")

# Analyze pattern types
# Pattern 1: exists + param_value = 'X'
# Pattern 2: exists + param_value = 'X' OR not exists (default fallback)
# Pattern 3: Multiple exists for different param_codes
# Pattern 4: CASE WHEN with #[PARAM]

pattern_a = 0  # exists(param_code='X' and param_value='Y')
pattern_b = 0  # exists(param_code='X' and param_value='Y') OR not exists(param_code='X')
pattern_c = 0  # Multiple different param_codes
pattern_d = 0  # #[PARAM] in expression (no URP_PARAM_CONFIG)

for item in all_param_sqls:
    sql_lower = item['sql'].lower()
    
    if '#[' in item['sql'] and 'URP_PARAM_CONFIG' not in item['sql']:
        pattern_d += 1
    elif 'not exists' in sql_lower and 'urp_param_config' in sql_lower:
        pattern_b += 1
    elif len(item['param_codes']) > 1:
        pattern_c += 1
    else:
        pattern_a += 1

print(f"\nPattern distribution:")
print(f"  Pattern A (exists param_value='X'): {pattern_a}")
print(f"  Pattern B (exists OR not exists - default fallback): {pattern_b}")
print(f"  Pattern C (multiple param_codes): {pattern_c}")
print(f"  Pattern D (#[PARAM] placeholder, no URP_PARAM_CONFIG): {pattern_d}")

# Show examples of each pattern
print("\n=== Pattern B examples (most common for parameter expansion) ===")
count = 0
for item in all_param_sqls:
    sql_lower = item['sql'].lower()
    if 'not exists' in sql_lower and 'urp_param_config' in sql_lower:
        # Extract the URP_PARAM_CONFIG block
        matches = re.finditer(
            r"(exists\s*\(\s*select[^)]+urp_param_config[^)]+\)|not exists\s*\(\s*select[^)]+urp_param_config[^)]+\))",
            item['sql'],
            re.IGNORECASE
        )
        blocks = [m.group(0) for m in matches]
        print(f"\n{item['alias']} ({item['check_code']}):")
        print(f"  param_codes: {item['param_codes']}")
        for b in blocks[:3]:
            print(f"  BLOCK: {b[:200]}")
        count += 1
        if count >= 5:
            break

print("\n=== Pattern D examples (#[PARAM] placeholder) ===")
count = 0
for item in all_param_sqls:
    if '#[' in item['sql'] and 'URP_PARAM_CONFIG' not in item['sql']:
        print(f"\n{item['alias']} ({item['check_code']}):")
        # Find all #[PARAM] references
        params = re.findall(r'#\[([A-Z_0-9]+)\]', item['sql'])
        print(f"  #[PARAM] refs: {params}")
        # Show a snippet around the #[PARAM]
        for p in set(params):
            idx = item['sql'].upper().find(f'#[{p}]')
            if idx >= 0:
                start = max(0, idx - 50)
                end = min(len(item['sql']), idx + len(p) + 60)
                print(f"  #[{p}] context: ...{item['sql'][start:end]}...")
        count += 1
        if count >= 3:
            break

# Now: how many params in code_param_list exist vs not in inspection SQL
import openpyxl as opx
wb3 = opx.load_workbook('F:/09_opencode/docs/code_param_list.xlsx', read_only=True, data_only=True)
ws3 = wb3[wb3.sheetnames[0]]
param_codes_in_config = set()
for row in ws3.iter_rows(min_row=2, values_only=True):
    if row[0]:
        param_codes_in_config.add(str(row[0]).strip())
wb3.close()

param_codes_in_templates = set()
for item in all_param_sqls:
    for p in item['param_codes']:
        param_codes_in_templates.add(p)

print(f"\n=== Param coverage ===")
print(f"Params in code_param_list.xlsx: {len(param_codes_in_config)}")
print(f"Params in inspection SQL templates: {len(param_codes_in_templates)}")
print(f"Params in templates but NOT in config: {param_codes_in_templates - param_codes_in_config}")