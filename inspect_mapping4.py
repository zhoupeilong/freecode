"""详细分析匹配缺口：哪些report_check_list报表名无法匹配到巡检列"""
import openpyxl
import os

_orig_dv_init = openpyxl.worksheet.datavalidation.DataValidation.__init__
def _patched_dv_init(self, *args, **kwargs):
    kwargs.pop('id', None)
    _orig_dv_init(self, *args, **kwargs)
openpyxl.worksheet.datavalidation.DataValidation.__init__ = _patched_dv_init

# 读取report_check_list.xlsx
check_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'report_check_list.xlsx')
wb2 = openpyxl.load_workbook(check_path, read_only=True)
check_tables = set()
for ws in wb2.worksheets:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) > 1 and row[1]:
            check_tables.add(str(row[1]).strip())
wb2.close()

# 读取巡检列sheet
xlsx_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'my_poor_solution.xlsx')
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb[wb.sheetnames[0]]

# 构建字段别名中的报表名集合
field_alias_tables = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or row[0] is None:
        continue
    field_alias = str(row[1]).strip() if row[1] else ""
    if '.' in field_alias:
        table_part = field_alias.split('.')[0]
        field_alias_tables.add(table_part)

# 分析匹配缺口
print("=== 匹配缺口分析 ===")
print(f"report_check_list 报表名总数: {len(check_tables)}")
print(f"巡检列 字段别名中的报表名: {len(field_alias_tables)}")
print(f"交集: {len(check_tables & field_alias_tables)}")

# 不匹配的报表名
unmatched = check_tables - field_alias_tables
matched = check_tables & field_alias_tables

print(f"\n=== 不匹配的报表名 ({len(unmatched)}) ===")
for t in sorted(unmatched)[:30]:
    print(f"  {t}")
if len(unmatched) > 30:
    print(f"  ... 共 {len(unmatched)} 个")

print(f"\n=== 匹配的报表名 ({len(matched)}) ===")
for t in sorted(matched)[:10]:
    print(f"  {t}")

# 关键发现：有些报表名是DM_开头的（如DM_COMM_ASSET_D_DTL_BAS），这些在巡检列中是什么格式？
print("\n=== DM_ 开头的报表名 ===")
dm_tables = [t for t in check_tables if t.startswith('DM_')]
for t in dm_tables[:10]:
    print(f"  {t}")

# 检查DM字段映射sheet能否帮助匹配
ws4 = wb[wb.sheetnames[3]]  # 关联到DM字段映射
dm_field_map_tables = set()
for row in ws4.iter_rows(min_row=2, values_only=True):
    if row and row[1]:
        dm_field_map_tables.add(str(row[1]).strip())

print(f"\nDM字段映射sheet中的报表名: {len(dm_field_map_tables)}")
print(f"  前10: {sorted(list(dm_field_map_tables))[:10]}")

# 检查DM字段映射sheet的匹配
dm_match = check_tables & dm_field_map_tables
print(f"  与report_check_list交集: {len(dm_match)}")
unmatch_dm = check_tables - dm_field_map_tables
print(f"  不匹配: {len(unmatch_dm)}")
for t in sorted(unmatch_dm)[:20]:
    print(f"    {t}")

wb.close()