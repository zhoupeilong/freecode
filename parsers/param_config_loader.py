"""
param_config_loader.py - 参数配置加载、分类与展开（v0.7 新增）

功能：
1. 从 code_param_list.xlsx 加载参数定义
2. 将参数分类为 business（业务参数，按值展开）和 personalization（个性化参数，保留URP_PARAM_CONFIG）
3. 计算业务参数的笛卡尔积展开
4. 从巡检列SQL模板和ETL清洗代码中提取参数引用

参数分类规则：
- 值数量 >= 10 且值列表包含机构代码/公司缩写 → personalization
- 参数名包含 GXHCS/ZXDJGID/JGMC 等机构相关关键字 → personalization
- 其余 → business
"""

import os
import re
import json
import openpyxl
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional
from itertools import product

# Patch openpyxl DataValidation
try:
    _orig_dv_init = openpyxl.worksheet.datavalidation.DataValidation.__init__
    def _patched_dv_init(self, *args, **kwargs):
        kwargs.pop('id', None)
        _orig_dv_init(self, *args, **kwargs)
    openpyxl.worksheet.datavalidation.DataValidation.__init__ = _patched_dv_init
except Exception:
    pass


@dataclass
class ParamDefinition:
    """参数定义"""
    param_code: str                          # 参数代码，如 CTRC_XMJLHXMFZRQSLJ
    param_name: str                          # 参数名称
    param_values: List[Tuple[str, str]]     # [(value, label), ...] 值列表
    default_value: str                       # 默认值
    category: str = ''                       # 'business' 或 'personalization'

    @property
    def value_count(self) -> int:
        return len(self.param_values)

    @property
    def value_list(self) -> List[str]:
        """返回纯值列表"""
        return [v for v, l in self.param_values]


@dataclass
class ParamFieldMapping:
    """字段→参数的映射关系"""
    dm_table: str            # DM表名
    dm_field: str            # DM字段名
    report_table: str        # URP报表名（可选）
    report_field: str        # URP报表字段名（可选）
    param_code: str          # 参数代码
    category: str            # 'business' 或 'personalization'
    case_when_expr: str      # CASE WHEN表达式摘要
    ref_source: str          # 参考来源（ETL文件名或巡检列）
    param_values: List[str] = field(default_factory=list)  # 参数有效值列表


@dataclass
class ParamExpansion:
    """一条参数展开记录"""
    check_code: str                          # 勾稽代码
    report_table: str                        # 报表名
    report_field: str                        # 报表字段名
    original_suc: str                        # 原始SUC编号
    business_params: Dict[str, str]          # {param_code: value} 业务参数值组合
    personal_params: List[str]               # 个性化参数代码列表（保留在SQL中）


def load_param_config(xlsx_path: str, logger=None) -> Dict[str, ParamDefinition]:
    """
    加载 code_param_list.xlsx，返回 {param_code: ParamDefinition}。
    
    Args:
        xlsx_path: code_param_list.xlsx 路径
        logger: Logger 实例
        
    Returns:
        dict: {param_code: ParamDefinition}
    """
    if not os.path.exists(xlsx_path):
        if logger:
            logger.warn(f"参数配置文件不存在: {xlsx_path}")
        return {}
    
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    
    result = {}
    row_count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        
        param_code = str(row[0]).strip()
        param_name = str(row[1]).strip() if row[1] else ''
        param_value = str(row[2]).strip() if row[2] else ''
        definition = str(row[3]).strip() if row[3] else ''
        default_value = str(row[4]).strip() if row[4] else ''
        
        # 解析参数定义JSON
        values = []
        if definition:
            try:
                defs = json.loads(definition)
                values = [(str(d.get('value', '')), str(d.get('label', ''))) for d in defs]
            except (json.JSONDecodeError, TypeError):
                # JSON解析失败，尝试简单解析
                pass
        
        # 分类
        category = classify_param(param_code, param_name, values)
        
        param_def = ParamDefinition(
            param_code=param_code,
            param_name=param_name,
            param_values=values,
            default_value=default_value,
            category=category,
        )
        
        # 处理同一param_code可能有多个行（不同值）
        # 合并值列表
        if param_code in result:
            existing = result[param_code]
            # 合并值（去重）
            existing_values = set(v for v, l in existing.param_values)
            for v, l in values:
                if v not in existing_values:
                    existing.param_values.append((v, l))
                    existing_values.add(v)
        else:
            result[param_code] = param_def
        
        row_count += 1
    
    wb.close()
    
    if logger:
        business_count = sum(1 for p in result.values() if p.category == 'business')
        personal_count = sum(1 for p in result.values() if p.category == 'personalization')
        logger.info(f"加载参数配置: {row_count} 行, {len(result)} 个唯一参数")
        logger.info(f"  业务参数: {business_count} 个")
        logger.info(f"  个性化参数: {personal_count} 个")
    
    return result


def classify_param(param_code: str, param_name: str, 
                   values: List[Tuple[str, str]]) -> str:
    """
    分类参数为 business 或 personalization。
    
    判定规则：
    1. 值数量 >= 10 且值列表中包含机构/公司关键字 → personalization
    2. 参数代码包含特定关键字 → personalization
    3. 其余 → business
    """
    # 个性化参数关键字（机构/公司级别）
    personalization_keywords = [
        'GXHCS',      # 个性化参数（66个信托公司）
        'ZXDJGID',     # 机构ID
        'JGMC',        # 机构名称
        'XTGSZXDBM',   # 信托公司中信登编码（68个值）
    ]
    
    # 检查参数代码
    param_upper = param_code.upper()
    for kw in personalization_keywords:
        if kw in param_upper:
            return 'personalization'
    
    # 检查值数量和内容
    n_values = len(values)
    if n_values >= 10:
        # 检查值列表中是否包含机构代码
        value_str = ' '.join(v + ' ' + l for v, l in values)
        org_keywords = ['XT', '信托', '公司', '编码', 'JGMC', '机构']
        for kw in org_keywords:
            if kw in value_str:
                return 'personalization'
        # 值数量多但没有机构关键字 → 可能是业务参数但需要特别关注
        # 如 TPRT_HTSFJZYJBJJZQSLJ 有11个值但仍是业务参数
    
    # 当前已知个性化参数精确匹配
    known_personal = {
        'TSIS_GXHCS',       # 66个信托公司
        'TUSP_XTGSZXDBM',   # 68个编码
        'TPRT_QLDJJGMCQSLJ', # 权利登记机构名称（取数逻辑，含多机构）
    }
    if param_code in known_personal:
        return 'personalization'
    
    return 'business'


def expand_business_params(
    param_config: Dict[str, ParamDefinition],
    business_param_codes: List[str]
) -> List[Dict[str, str]]:
    """
    计算业务参数的笛卡尔积，返回所有参数值组合。
    
    Args:
        param_config: 参数配置字典
        business_param_codes: 需要展开的业务参数代码列表
        
    Returns:
        list: [{param_code: value, ...}, ...] 每个元素代表一组参数值组合
    """
    if not business_param_codes:
        return [{}]  # 无参数，返回一个空组合
    
    # 收集每个业务参数的值列表
    param_value_lists = []
    param_codes_ordered = []
    
    for code in business_param_codes:
        if code in param_config:
            param = param_config[code]
            if param.param_values:
                param_value_lists.append(param.value_list)
                param_codes_ordered.append(code)
            else:
                # 参数定义中没有值列表，使用默认值
                if param.default_value:
                    param_value_lists.append([param.default_value])
                    param_codes_ordered.append(code)
                else:
                    # 既没有值列表也没有默认值，跳过此参数
                    if code not in param_config:
                        # 未知参数，使用占位值
                        param_value_lists.append(['UNKNOWN'])
                        param_codes_ordered.append(code)
        else:
            # 参数不在配置中（如 SRDT5_YGH 等价于 CTRC_YGH）
            # 尝试等价映射
            equiv_code = _get_equivalent_param(code)
            if equiv_code and equiv_code in param_config:
                param = param_config[equiv_code]
                if param.param_values:
                    param_value_lists.append(param.value_list)
                    param_codes_ordered.append(code)
                elif param.default_value:
                    param_value_lists.append([param.default_value])
                    param_codes_ordered.append(code)
            else:
                # 未知参数，标记单值
                param_value_lists.append(['1'])
                param_codes_ordered.append(code)
    
    if not param_value_lists:
        return [{}]
    
    # 笛卡尔积
    combos = list(product(*param_value_lists))
    
    result = []
    for combo in combos:
        result.append(dict(zip(param_codes_ordered, combo)))
    
    return result


# 参数代码等价映射
# ETL清洗代码和巡检列SQL中可能使用不同的参数代码名
_PARAM_EQUIVALENCE = {
    'SRDT5_YGH': 'CTRC_YGH',           # 员工号取数场景
    'SRDT5_YGXXBQSLY': 'CTRC_YGXXBQSLY', # 员工信息指标表取数来源
    'SRDT5_GLFBS': 'CTRC_GLFBS',         # 关联方标识
    'TPRT_CWXXQSLY': 'TPRP_CWXXQSLY',     # 全流程_财务信息取数来源
}


def _get_equivalent_param(param_code: str) -> Optional[str]:
    """获取参数的等价代码"""
    return _PARAM_EQUIVALENCE.get(param_code)


def extract_params_from_sql(sql: str) -> Tuple[List[str], List[str]]:
    """
    从SQL模板中提取参数引用。
    
    支持两种模式：
    1. URP_PARAM_CONFIG 标量子查询: param_code = 'XXX'
    2. #[PARAM] 占位符: #[CTRC_YGH]
    
    Returns:
        (param_codes_from_config, param_codes_from_hash)
        - param_codes_from_config: URP_PARAM_CONFIG中的参数代码列表
        - param_codes_from_hash: #[PARAM]占位符中的参数代码列表
    """
    # 提取 URP_PARAM_CONFIG 引用
    config_pattern = re.compile(
        r"param_code\s*=\s*'([^']+)'",
        re.IGNORECASE
    )
    config_codes = list(set(config_pattern.findall(sql)))
    
    # 提取 #[PARAM] 占位符
    hash_pattern = re.compile(r"#\[([A-Z_0-9]+)\]", re.IGNORECASE)
    hash_codes = list(set(hash_pattern.findall(sql)))
    
    return config_codes, hash_codes


def get_all_param_codes(sql: str) -> List[str]:
    """获取SQL中所有引用的参数代码（去重）"""
    config_codes, hash_codes = extract_params_from_sql(sql)
    return list(set(config_codes + hash_codes))


def classify_params_in_sql(
    sql: str,
    param_config: Dict[str, ParamDefinition]
) -> Tuple[List[str], List[str]]:
    """
    将SQL中的参数引用分类为 business 和 personalization。
    
    Returns:
        (business_codes, personal_codes)
    """
    all_codes = get_all_param_codes(sql)
    business_codes = []
    personal_codes = []
    
    for code in all_codes:
        # 查找参数定义（考虑等价映射）
        actual_code = code
        if code not in param_config:
            equiv = _get_equivalent_param(code)
            if equiv and equiv in param_config:
                actual_code = equiv
        
        if actual_code in param_config:
            category = param_config[actual_code].category
        else:
            # 不在配置中的参数，根据关键字判断
            category = classify_param(code, '', [])
        
        if category == 'business':
            business_codes.append(code)
        else:
            personal_codes.append(code)
    
    return business_codes, personal_codes


def build_param_value_label(
    param_code: str,
    value: str,
    param_config: Dict[str, ParamDefinition]
) -> str:
    """构建参数值标签，如 'CTRC_XMJLHXMFZRQSLJ=1(项目经理)'"""
    # 查找参数定义（考虑等价映射）
    actual_code = param_code
    if param_code not in param_config:
        equiv = _get_equivalent_param(param_code)
        if equiv and equiv in param_config:
            actual_code = equiv
    
    if actual_code in param_config:
        param = param_config[actual_code]
        for v, l in param.param_values:
            if v == value:
                short_label = l[:20] if len(l) > 20 else l
                return f"{param_code}={value}({short_label})"
    
    return f"{param_code}={value}"


def save_param_config_report(
    param_config: Dict[str, ParamDefinition],
    output_path: str
):
    """保存参数配置分类报告"""
    import csv
    
    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow([
            'param_code', 'param_name', 'category', 'value_count', 
            'default_value', 'values', 'labels'
        ])
        
        for code, param in sorted(param_config.items()):
            values_str = '|'.join(v for v, l in param.param_values)
            labels_str = '|'.join(l[:30] for v, l in param.param_values)
            writer.writerow([
                param.param_code,
                param.param_name,
                param.category,
                param.value_count,
                param.default_value,
                values_str,
                labels_str
            ])


if __name__ == '__main__':
    # 测试入口
    xlsx_path = os.path.abspath(os.path.join(
        os.path.dirname(__file__), '..', 'code_param_list.xlsx'))
    
    from utils.logger import Logger
    test_log = os.path.abspath(os.path.join(
        os.path.dirname(__file__), '..', 'output_stg_file', 'test_param_log.txt'))
    os.makedirs(os.path.dirname(test_log), exist_ok=True)
    logger = Logger(test_log)
    
    param_config = load_param_config(xlsx_path, logger)
    
    print(f"\n=== 参数配置测试 ===")
    print(f"总参数数: {len(param_config)}")
    
    business_params = {k: v for k, v in param_config.items() if v.category == 'business'}
    personal_params = {k: v for k, v in param_config.items() if v.category == 'personalization'}
    
    print(f"业务参数: {len(business_params)} 个")
    print(f"个性化参数: {len(personal_params)} 个")
    
    # 测试关键参数
    test_codes = ['CTRC_XMJLHXMFZRQSLJ', 'CTRC_YGH', 'TSIS_GXHCS']
    for code in test_codes:
        equiv = _get_equivalent_param(code) or code
        if equiv in param_config:
            p = param_config[equiv]
            print(f"\n  {code}:")
            print(f"    名称: {p.param_name}")
            print(f"    分类: {p.category}")
            print(f"    值数量: {p.value_count}")
            print(f"    默认值: {p.default_value}")
            for v, l in p.param_values[:5]:
                print(f"    {v}: {l}")
            if p.value_count > 5:
                print(f"    ... 还有 {p.value_count - 5} 个值")
    
    # 测试展开
    print(f"\n=== 笛卡尔积展开测试 ===")
    business_codes = ['CTRC_XMJLHXMFZRQSLJ', 'CTRC_YGH']
    combos = expand_business_params(param_config, business_codes)
    print(f"CTRC_XMJLHXMFZRQSLJ × CTRC_YGH = {len(combos)} 个组合:")
    for combo in combos:
        labels = []
        for code, value in combo.items():
            labels.append(build_param_value_label(code, value, param_config))
        print(f"  {', '.join(labels)}")
    
    # 保存分类报告
    report_path = os.path.abspath(os.path.join(
        os.path.dirname(__file__), '..', 'output_stg_file', 'param_config_report.csv'))
    save_param_config_report(param_config, report_path)
    print(f"\n参数分类报告已保存: {report_path}")