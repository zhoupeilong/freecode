"""
report_check_parser.py - 读取 report_check_list.xlsx，生成 ReportCheckInfo 列表

输入：report_check_list.xlsx
    列结构：
    A: ID          - 行记录ID，格式为 "报表名.字段名"
    B: 报表名称     - 如 URP_SRDT5_GDXXB
    C: 报表字段     - 如 ID
    D: 勾稽代码     - 如 102.1679 或 DQ03.A0371
    E: 报表勾稽名称 - 如 <股东信息表-ID>不应为空
    F: 是否必填     - "必填" 或 "条件性必填"
    G: 前置条件     - 如 <是否互联网贷款>选择是时
    H: 前置表       - 如 URP3_TUSP.DM_CTRC_CPTZZBB
    I: 前置字段     - 如 sfhlwdk
    J: 前置字段值   - 如 '1'

输出：List[ReportCheckInfo]
"""

from dataclasses import dataclass, field
from typing import Optional, List
import openpyxl
import os
import re


@dataclass
class ReportCheckInfo:
    """报表勾稽检查项的数据模型"""
    record_id: str                # A列: ID，如 "URP_SRDT5_GDXXB.ID"
    report_table: str             # B列: 报表名称，如 "URP_SRDT5_GDXXB"
    report_field: str             # C列: 报表字段，如 "ID"
    check_code: str               # D列: 勾稽代码，如 "102.1679" 或 "DQ03.A0371"
    check_name: str               # E列: 勾稽名称，如 "<股东信息表-ID>不应为空"
    is_mandatory: str             # F列: "必填" 或 "条件性必填"
    precondition: Optional[str]   # G列: 前置条件描述，如 "<是否互联网贷款>选择是时"
    pre_table: Optional[str]      # H列: 前置表，如 "URP3_TUSP.DM_CTRC_CPTZZBB"
    pre_column: Optional[str]    # I列: 前置字段，如 "sfhlwdk"
    pre_value: Optional[str]     # J列: 前置字段值，如 "'1'"

    @property
    def is_conditional(self) -> bool:
        """是否为条件必填"""
        return "条件" in self.is_mandatory

    @property
    def has_precondition(self) -> bool:
        """是否有前置条件"""
        return self.pre_table is not None or self.pre_column is not None


def load_report_checks(xlsx_path: str) -> List[ReportCheckInfo]:
    """
    读取 report_check_list.xlsx，返回 ReportCheckInfo 列表。

    Args:
        xlsx_path: Excel 文件的完整路径

    Returns:
        List[ReportCheckInfo]: 所有检查项的列表

    Raises:
        FileNotFoundError: 如果文件不存在
        ValueError: 如果文件格式不符合预期
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"文件不存在: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    # 查找包含检查数据的 sheet（通常是 Sheet1）
    ws = None
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # 检查第一行是否包含预期列头
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row and len(first_row) >= 6:
            ws = sheet
            break

    if ws is None:
        raise ValueError("未找到包含检查数据的 sheet")

    checks = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # 跳过空行
        if not row or row[0] is None:
            continue

        # 确保至少有 6 列
        if len(row) < 6:
            skipped += 1
            continue

        record_id = str(row[0]).strip() if row[0] else ""
        report_table = str(row[1]).strip() if row[1] else ""
        report_field = str(row[2]).strip() if row[2] else ""
        check_code = str(row[3]).strip() if row[3] else ""
        check_name = str(row[4]).strip() if row[4] else ""
        is_mandatory = str(row[5]).strip() if row[5] else ""

        # 前置条件列（G-J），可能为空
        precondition = str(row[6]).strip() if len(row) > 6 and row[6] else None
        pre_table = str(row[7]).strip() if len(row) > 7 and row[7] else None
        pre_column = str(row[8]).strip() if len(row) > 8 and row[8] else None
        pre_value = str(row[9]).strip() if len(row) > 9 and row[9] else None

        # 跳过无效行
        if not record_id or not check_code:
            skipped += 1
            continue

        info = ReportCheckInfo(
            record_id=record_id,
            report_table=report_table,
            report_field=report_field,
            check_code=check_code,
            check_name=check_name,
            is_mandatory=is_mandatory,
            precondition=precondition,
            pre_table=pre_table,
            pre_column=pre_column,
            pre_value=pre_value,
        )
        checks.append(info)

    wb.close()

    print(f"[report_check_parser] 读取完成: 共 {len(checks)} 条检查项, 跳过 {skipped} 条无效行")

    # 统计
    mandatory_count = sum(1 for c in checks if not c.is_conditional)
    conditional_count = sum(1 for c in checks if c.is_conditional)
    has_pre_count = sum(1 for c in checks if c.has_precondition)

    print(f"[report_check_parser] 统计: 必填 {mandatory_count} 条, 条件必填 {conditional_count} 条, 有前置条件 {has_pre_count} 条")

    return checks


if __name__ == "__main__":
    # 测试入口
    xlsx_path = os.path.join(os.path.dirname(__file__), "..", "..", "report_check_list.xlsx")
    xlsx_path = os.path.abspath(xlsx_path)

    checks = load_report_checks(xlsx_path)

    # 打印前5条
    print("\n=== 前5条检查项 ===")
    for i, c in enumerate(checks[:5]):
        print(f"\n--- 检查项 {i+1} ---")
        print(f"  ID: {c.record_id}")
        print(f"  报表: {c.report_table}.{c.report_field}")
        print(f"  勾稽代码: {c.check_code}")
        print(f"  勾稽名称: {c.check_name}")
        print(f"  是否必填: {c.is_mandatory} (条件必填={c.is_conditional})")
        print(f"  前置条件: {c.precondition}")
        print(f"  前置表: {c.pre_table}")
        print(f"  前置字段: {c.pre_column}")
        print(f"  前置字段值: {c.pre_value}")

    # 打印几条条件必填
    print("\n=== 条件必填示例（前3条）===")
    for i, c in enumerate([c for c in checks if c.is_conditional][:3]):
        print(f"\n--- 条件必填 {i+1} ---")
        print(f"  ID: {c.record_id}")
        print(f"  勾稽名称: {c.check_name}")
        print(f"  前置条件: {c.precondition}")
        print(f"  前置表: {c.pre_table}")
        print(f"  前置字段: {c.pre_column}")
        print(f"  前置字段值: {c.pre_value}")