"""
etl_mapper.py - 从 my_poor_solution.xlsx 多 sheet 提取 STG→DW→DM 映射

映射策略（优先级从高到低）：
1. 巡检列 sheet（主映射源）：通过 字段别名(报表名.字段名) → 匹配 report_check_list 的 report_table.report_field
   - 提供：DM表、DM字段、STG表、STG字段、关联体系(SQL模板)
2. 关联到DM字段映射 sheet：作为补充映射
3. STG到DM表映射 sheet：SQL模板与关联体系

核心变更：不再通过 DM 表名模糊匹配，而是通过 字段别名 精确匹配。
"""

import os
import re
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Tuple
import openpyxl

# Patch openpyxl DataValidation to handle 'id' parameter in newer xlsx files
try:
    _orig_dv_init = openpyxl.worksheet.datavalidation.DataValidation.__init__
    def _patched_dv_init(self, *args, **kwargs):
        kwargs.pop('id', None)
        _orig_dv_init(self, *args, **kwargs)
    openpyxl.worksheet.datavalidation.DataValidation.__init__ = _patched_dv_init
except Exception:
    pass


@dataclass
class InspectionRow:
    """巡检列 sheet 中的一行数据"""
    row_id: int                      # 字段序号
    field_alias: str                  # 字段别名（如 URP_TPRT_CPJBXX.XTDJXTCPBM）
    field_desc: str                  # 字段说明
    check_code: str                  # 勾稽代码（如 DQ01.A0001）
    check_name: str                  # 勾稽名称
    is_mandatory: str                # 是否必填
    precondition: str                # 前置条件
    pre_table: str                   # 前置表
    pre_column: str                  # 前置字段
    pre_value: str                   # 前置字段值
    extra_condition: str             # 额外条件
    dm_table: str                    # DM表
    dm_field: str                    # DM字段
    check_expression: str            # 检验表达式
    stg_field_count: int             # STG字段涉及层数
    pre_condition_count: int         # 前置条件层数
    stg_table_cn: str                # STG表中文名
    stg_field_cn: str                # STG字段中文名
    stg_table: str                   # 取数来源（STG表名）
    stg_field: str                   # 取数来源字段
    link_id: str                     # 关联体系ID
    sql_template: str                # 关联内容（SQL模板）
    suc_number: str                  # 编号（如 SUC0001）
    is_from_dm_supplement: bool = False  # 是否来自DM字段映射补充（非巡检列精确匹配）

    @property
    def report_table(self) -> str:
        """从字段别名提取报表名"""
        if '.' in self.field_alias:
            return self.field_alias.split('.')[0]
        return ''

    @property
    def report_field(self) -> str:
        """从字段别名提取字段名"""
        if '.' in self.field_alias:
            return self.field_alias.split('.')[1]
        return ''


@dataclass
class DMFieldMapping:
    """关联到DM字段映射 sheet 中的一行数据"""
    id: str                # 如 URP_SRDT5_CPJBXXB.CCLY
    tab_name: str          # 如 URP_SRDT5_CPJBXXB
    tab_col: str           # 如 CCLY
    ref_tab: str           # 如 DM_CTRC_CPJBXXZBB
    ref_col: str           # 如 CCLY
    source_table: str      # 取数来源（STG表）
    source_col: str        # 取数来源字段
    check_mode: str        # 如 required


@dataclass
class STGDWMapping:
    """STG → DW → DM 的关联映射（来自 STG到DM表映射 sheet）"""
    stg_table: str           # STG 表名
    dm_table: str            # DM 表名
    dw_tables: List[str]     # 中间DW表列表
    on_dm_dw: str            # DM → DW 的关联条件
    on_dw_stg: str           # DW → STG 的关联条件
    stg_key: str             # STG 业务主键表达式
    sql_template: str         # 已有的SQL模板
    data_source: str = ""    # 数据来源限制
    where_clause: str = ""   # 外层 WHERE 条件


@dataclass
class ETLMapping:
    """完整的 ETL 层级映射"""
    # 巡检列映射：字段别名 → InspectionRow列表
    # 一个字段别名可能有多行（同一字段对应多个STG来源）
    inspection_map: Dict[str, List[InspectionRow]] = field(default_factory=dict)
    # DM字段映射：TAB_NAME.COL → DMFieldMapping列表
    dm_field_map: Dict[str, List[DMFieldMapping]] = field(default_factory=dict)
    # STG→DM映射：关联体系ID → STGDWMapping
    stg_dw_map: Dict[str, STGDWMapping] = field(default_factory=dict)
    # STG主键映射：STG表名 → 主键表达式
    stg_key_map: Dict[str, str] = field(default_factory=dict)


def load_inspection_sheet(xlsx_path: str, logger=None) -> Dict[str, List[InspectionRow]]:
    """
    加载巡检列 sheet，构建 字段别名(报表名.字段名) → InspectionRow 的映射。

    Args:
        xlsx_path: my_poor_solution.xlsx 路径
        logger: Logger 实例

    Returns:
        dict: {字段别名: [InspectionRow, ...]}
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]  # 巡检列 sheet

    result = {}
    row_count = 0
    has_sql = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        def _safe_str(val, default=''):
            if val is None or str(val).strip() in ('0', 'None', ''):
                return default
            return str(val).strip()

        field_alias = _safe_str(row[1])
        if not field_alias or '.' not in field_alias:
            continue

        sql_template = _safe_str(row[22]) if len(row) > 22 else ''

        insp_row = InspectionRow(
            row_id=int(row[0]) if row[0] else 0,
            field_alias=field_alias,
            field_desc=_safe_str(row[2]),
            check_code=_safe_str(row[3]),
            check_name=_safe_str(row[4]),
            is_mandatory=_safe_str(row[5]),
            precondition=_safe_str(row[6]),
            pre_table=_safe_str(row[7]),
            pre_column=_safe_str(row[8]),
            pre_value=_safe_str(row[9]),
            extra_condition=_safe_str(row[10]),
            dm_table=_safe_str(row[12]),
            dm_field=_safe_str(row[13]),
            check_expression=_safe_str(row[14]),
            stg_field_count=int(row[15]) if row[15] and str(row[15]).isdigit() else 0,
            pre_condition_count=int(row[16]) if row[16] and str(row[16]).isdigit() else 0,
            stg_table_cn=_safe_str(row[17]),
            stg_field_cn=_safe_str(row[18]),
            stg_table=_safe_str(row[19]),
            stg_field=_safe_str(row[20]),
            link_id=_safe_str(row[21]),
            sql_template=sql_template,
            suc_number=_safe_str(row[28]),
        )

        if field_alias not in result:
            result[field_alias] = []
        result[field_alias].append(insp_row)
        row_count += 1

        if sql_template:
            has_sql += 1

    wb.close()

    if logger:
        logger.info(f"加载巡检列 sheet: {row_count} 行数据, {has_sql} 行有SQL模板")

    return result


def load_dm_field_mapping(xlsx_path: str, logger=None) -> Dict[str, List[DMFieldMapping]]:
    """
    加载 关联到DM字段映射 sheet，构建 TAB_NAME.COL → DMFieldMapping 的映射。

    Args:
        xlsx_path: my_poor_solution.xlsx 路径
        logger: Logger 实例

    Returns:
        dict: {TAB_NAME.COL: [DMFieldMapping, ...]}
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[3]]  # 关联到DM字段映射 sheet

    result = {}
    row_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue

        def _safe_str(val, default=''):
            if val is None:
                return default
            return str(val).strip()

        dm_id = _safe_str(row[0])
        tab_name = _safe_str(row[1])
        tab_col = _safe_str(row[2])
        ref_tab = _safe_str(row[3])
        ref_col = _safe_str(row[4])
        source_table = _safe_str(row[5])
        source_col = _safe_str(row[6])
        check_mode = _safe_str(row[7])

        if not tab_name or not ref_tab:
            continue

        mapping = DMFieldMapping(
            id=dm_id,
            tab_name=tab_name,
            tab_col=tab_col,
            ref_tab=ref_tab,
            ref_col=ref_col,
            source_table=source_table,
            source_col=source_col,
            check_mode=check_mode,
        )

        key = f"{tab_name}.{tab_col}"
        if key not in result:
            result[key] = []
        result[key].append(mapping)
        row_count += 1

    wb.close()

    if logger:
        logger.info(f"加载DM字段映射: {row_count} 行数据, {len(result)} 个唯一键")

    return result


def load_stg_dw_mappings(xlsx_path: str, logger=None) -> Dict[str, STGDWMapping]:
    """
    加载 STG到DM表映射 sheet，构建 关联体系ID → STGDWMapping 的映射。

    Args:
        xlsx_path: my_poor_solution.xlsx 路径
        logger: Logger 实例

    Returns:
        dict: {关联体系ID: STGDWMapping}
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[1]]  # STG到DM表映射 sheet

    # 先加载STG主键
    stg_key_map = {}
    stg_key_sheet = None
    for name in wb.sheetnames:
        if '主键' in name:
            stg_key_sheet = wb[name]
            break
    if stg_key_sheet:
        for row in stg_key_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                stg_key_map[str(row[0]).strip()] = str(row[1]).strip()

    result = {}
    row_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5 or not row[1]:
            continue

        def _safe_str(val, default=''):
            if val is None:
                return default
            return str(val).strip()

        stg_table = _safe_str(row[1])
        link_id = _safe_str(row[0])  # 关联体系ID
        stg_key_raw = _safe_str(row[2])
        dm_table = _safe_str(row[3])
        sql_template = _safe_str(row[4]) if len(row) > 4 else ''
        custom_on = _safe_str(row[5]) if len(row) > 5 else ''

        if not stg_table or not dm_table:
            continue

        # 解析SQL模板
        dw_tables, on_dm_dw, on_dw_stg, data_source, where_clause = _parse_sql_template(sql_template)

        stg_key = stg_key_map.get(stg_table, stg_key_raw)

        mapping = STGDWMapping(
            stg_table=stg_table,
            dm_table=dm_table,
            dw_tables=dw_tables,
            on_dm_dw=on_dm_dw,
            on_dw_stg=on_dw_stg,
            stg_key=stg_key,
            sql_template=sql_template,
            data_source=data_source,
            where_clause=where_clause,
        )

        # 用关联体系ID做key
        if link_id:
            result[link_id] = mapping
        result[f"{stg_table}.{dm_table}"] = mapping
        row_count += 1

    wb.close()

    if logger:
        logger.info(f"加载STG→DW映射: {row_count} 条, {len(result)} 个唯一键")

    return result


def _parse_sql_template(sql: str) -> Tuple[List[str], str, str, str, str]:
    """解析 SQL 模板中的 DW 表、关联条件等"""
    if not sql or sql.startswith('='):
        return [], "", "", "", ""

    dw_tables = []
    on_dm_dw = ""
    on_dw_stg = ""
    data_source = ""
    where_clause = ""

    join_pattern = re.compile(
        r'(?i)(?:INNER|LEFT)\s+JOIN\s+([\w.]+)\s+(\w+)?\s*\n?\s*ON\s+([^\n,]+)',
        re.IGNORECASE | re.DOTALL
    )

    for match in join_pattern.finditer(sql):
        table_name = match.group(1).strip()
        on_clause = match.group(3).strip()

        upper_table = table_name.upper()
        if 'DW_D_' in upper_table or 'dw_d_' in table_name.lower():
            dw_tables.append(table_name)
            if not on_dm_dw and 'DM_' in on_clause.upper():
                on_dm_dw = on_clause
            elif not on_dw_stg and ('T2_' in on_clause.upper() or 'TS_' in on_clause.upper()):
                on_dw_stg = on_clause

    ds_pattern = re.compile(r"(?i)data_source\s*=\s*'(\w+)'")
    ds_match = ds_pattern.search(sql)
    if ds_match:
        data_source = ds_match.group(1)

    return dw_tables, on_dm_dw, on_dw_stg, data_source, where_clause


def load_my_poor_solution(xlsx_path: str, logger=None) -> ETLMapping:
    """
    加载 my_poor_solution.xlsx 全部映射数据。

    返回的 ETLMapping 包含：
    - inspection_map: 巡检列 字段别名→InspectionRow 映射（主映射源）
    - dm_field_map: DM字段映射（补充）
    - stg_dw_map: STG→DW→DM 关联体系（SQL模板）
    - stg_key_map: STG主键表达式
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"文件不存在: {xlsx_path}")

    mapping = ETLMapping()

    # 1. 加载巡检列（主映射源）
    mapping.inspection_map = load_inspection_sheet(xlsx_path, logger)

    # 2. 加载DM字段映射（补充）
    mapping.dm_field_map = load_dm_field_mapping(xlsx_path, logger)

    # 3. 加载STG→DW映射（SQL模板）
    mapping.stg_dw_map = load_stg_dw_mappings(xlsx_path, logger)

    # 4. 加载STG主键
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    stg_key_sheet = None
    for name in wb.sheetnames:
        if '主键' in name:
            stg_key_sheet = wb[name]
            break
    if stg_key_sheet:
        for row in stg_key_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                mapping.stg_key_map[str(row[0]).strip()] = str(row[1]).strip()
    wb.close()

    # 统计
    if logger:
        total_rows = sum(len(v) for v in mapping.inspection_map.values())
        has_sql = sum(1 for rows in mapping.inspection_map.values() for r in rows if r.sql_template)
        has_stg = sum(1 for rows in mapping.inspection_map.values() for r in rows if r.stg_table)
        logger.info(f"映射统计:")
        logger.info(f"  巡检列: {len(mapping.inspection_map)} 个字段别名, {total_rows} 行数据")
        logger.info(f"  有SQL模板: {has_sql} 行, 有STG表: {has_stg} 行")
        logger.info(f"  DM字段映射: {len(mapping.dm_field_map)} 个键")
        logger.info(f"  STG→DW映射: {len(mapping.stg_dw_map)} 个键")
        logger.info(f"  STG主键: {len(mapping.stg_key_map)} 个")

    return mapping


def find_mapping_for_check(check_info, mapping: ETLMapping, logger=None) -> Optional[InspectionRow]:
    """
    为一条检查项查找最佳映射。

    匹配策略：
    1. 精确匹配：report_table.report_field → 巡检列字段别名
    2. 大小写不敏感匹配
    3. DM字段映射补充 - 使用DM字段映射表精确定位DM表和字段
    
    ⚠️ 不再使用"同报表名取第一行"策略，因为字段名不同时会导致SQL检查了错误的字段。
    
    Args:
        check_info: ReportCheckInfo 实例
        mapping: ETLMapping 实例
        logger: Logger 实例

    Returns:
        InspectionRow 或 None
    """
    report_table = check_info.report_table.strip()
    report_field = check_info.report_field.strip()

    # 1. 精确匹配：报表名.字段名
    field_alias = f"{report_table}.{report_field}"
    if field_alias in mapping.inspection_map:
        rows = mapping.inspection_map[field_alias]
        # 优先选择有SQL模板的行
        with_sql = [r for r in rows if r.sql_template]
        if with_sql:
            return with_sql[0]
        return rows[0]

    # 2. 尝试大小写不敏感匹配
    field_alias_upper = field_alias.upper()
    for key, rows in mapping.inspection_map.items():
        if key.upper() == field_alias_upper:
            with_sql = [r for r in rows if r.sql_template]
            if with_sql:
                return with_sql[0]
            return rows[0]

    # 3. DM字段映射补充 - 使用DM字段映射表精确定位
    dm_key = f"{report_table}.{report_field}"
    if dm_key in mapping.dm_field_map:
        dm_rows = mapping.dm_field_map[dm_key]
        if dm_rows:
            dm_row = dm_rows[0]

            # 4a. 先尝试在同一DM表中找到有SQL模板的巡检列参考行
            # 这样可以复用该DM表的SQL模板（替换字段即可）
            ref_row = _find_reference_row_for_dm_table(dm_row.ref_tab, mapping)
            if ref_row:
                # 创建基于参考行的映射，替换关键字段
                return InspectionRow(
                    row_id=0,
                    field_alias=field_alias,
                    field_desc=check_info.check_name,
                    check_code=check_info.check_code,
                    check_name=check_info.check_name,
                    is_mandatory=check_info.is_mandatory,
                    precondition=check_info.precondition or '',
                    pre_table=check_info.pre_table or '',
                    pre_column=check_info.pre_column or '',
                    pre_value=check_info.pre_value or '',
                    extra_condition='',
                    dm_table=dm_row.ref_tab,
                    dm_field=dm_row.ref_col or check_info.report_field,
                    check_expression=f"trim(k.{dm_row.ref_col or check_info.report_field}) IS NULL",
                    stg_field_count=0,
                    pre_condition_count=0,
                    stg_table_cn=ref_row.stg_table_cn if ref_row else '',
                    stg_field_cn='',
                    stg_table=ref_row.stg_table if ref_row else '',
                    stg_field=dm_row.source_col or check_info.report_field,
                    link_id=ref_row.link_id if ref_row else '',
                    sql_template=ref_row.sql_template if ref_row else '',
                    suc_number='',
                    is_from_dm_supplement=True,  # 标记为DM补充行
                )

            # 4b. 没有参考行，创建最基本信息
            return InspectionRow(
                row_id=0,
                field_alias=field_alias,
                field_desc=check_info.check_name,
                check_code=check_info.check_code,
                check_name=check_info.check_name,
                is_mandatory=check_info.is_mandatory,
                precondition=check_info.precondition or '',
                pre_table=check_info.pre_table or '',
                pre_column=check_info.pre_column or '',
                pre_value=check_info.pre_value or '',
                extra_condition='',
                dm_table=dm_row.ref_tab,
                dm_field=dm_row.ref_col or check_info.report_field,
                check_expression=f"trim(k.{dm_row.ref_col or check_info.report_field}) IS NULL",
                stg_field_count=0,
                pre_condition_count=0,
                stg_table_cn='',
                stg_field_cn='',
                stg_table=dm_row.source_table or '',
stg_field=dm_row.source_col or check_info.report_field,
                    link_id='',
                    sql_template='',
                    suc_number='',
                    is_from_dm_supplement=True,  # 标记为DM补充行
                )

    if logger:
        logger.warn(
            f"未找到映射: 报表={report_table}, 字段={report_field}, "
            f"别名={field_alias}",
            Logger.WARN_MAPPING_MISSING if logger else "WARN_MAPPING_MISSING"
        )

    return None


def _find_reference_row_for_dm_table(dm_table: str, mapping: 'ETLMapping') -> Optional[InspectionRow]:
    """
    在巡检列中查找同一DM表的有SQL模板的参考行。

    用于当精确匹配失败时，复用同DM表的SQL模板结构。
    """
    best_row = None
    for key, rows in mapping.inspection_map.items():
        for r in rows:
            if r.dm_table == dm_table and r.sql_template:
                return r  # 优先返回有SQL模板的
            if r.dm_table == dm_table and best_row is None:
                best_row = r
    return best_row


def get_stg_dw_mapping(link_id: str, mapping: ETLMapping) -> Optional[STGDWMapping]:
    """
    通过关联体系ID获取 STG→DW→DM 映射。

    Args:
        link_id: 关联体系ID（如 T2_TCMP_PROJECTINFO.DM_CTRC_CPJBXXZBB）
        mapping: ETLMapping 实例

    Returns:
        STGDWMapping 或 None
    """
    if link_id in mapping.stg_dw_map:
        return mapping.stg_dw_map[link_id]

    # 尝试不带schema前缀匹配
    for key, m in mapping.stg_dw_map.items():
        if key.endswith(f'.{link_id}') or link_id.endswith(f'.{key}'):
            return m

    return None


# 让 get_warn_type 兼容
class Logger:
    WARN_MAPPING_MISSING = "WARN_MAPPING_MISSING"


if __name__ == "__main__":
    # 测试入口
    xlsx_path = os.path.abspath(os.path.join(
        os.path.dirname(__file__), "..", "..", "my_poor_solution.xlsx"))

    from utils.logger import Logger as RealLogger
    test_log = os.path.abspath(os.path.join(
        os.path.dirname(__file__), "..", "..", "output_stg_file", "test_mapping_log.txt"))
    logger = RealLogger(test_log)

    mapping = load_my_poor_solution(xlsx_path, logger)

    print(f"\n=== 映射测试 ===")
    print(f"巡检列字段别名数: {len(mapping.inspection_map)}")
    print(f"DM字段映射键数: {len(mapping.dm_field_map)}")
    print(f"STG→DW映射数: {len(mapping.stg_dw_map)}")

    # 测试几个字段别名
    test_aliases = [
        'URP_TPRT_CPJBXX.XTDJXTCPBM',
        'URP_TPRT_CPJBXX.SYFS',
        'URP_SRDT5_GDXXB.ID',  # 可能不在巡检列
    ]

    for alias in test_aliases:
        if alias in mapping.inspection_map:
            rows = mapping.inspection_map[alias]
            r = rows[0]
            print(f"\n  {alias}:")
            print(f"    DM表: {r.dm_table}, DM字段: {r.dm_field}")
            print(f"    STG表: {r.stg_table}, STG字段: {r.stg_field}")
            print(f"    有SQL模板: {'是' if r.sql_template else '否'}")
            print(f"    关联体系ID: {r.link_id}")
        else:
            print(f"\n  {alias}: 未找到")