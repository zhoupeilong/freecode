"""
parse_etl_params.py - 从ETL清洗代码和巡检列SQL中提取字段→参数的依赖关系（v0.7 新增）

功能：
1. 扫描 etl_code_list/ 目录中的ETL清洗SQL文件，提取 #[PARAM] 占位符和 CASE WHEN 结构
2. 扫描巡检列SQL模板，提取 URP_PARAM_CONFIG 引用
3. 建立 DM字段→参数代码 的映射关系
4. 输出 param_field_mapping.xlsx

输出格式：
| DM_TABLE | DM_FIELD | PARAM_CODE | PARAM_CATEGORY | CASE_WHEN_EXPR | REF_SOURCE | PARAM_VALUES |
"""

import os
import re
import json
import glob
import openpyxl
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass, field

# Patch openpyxl
try:
    _orig_dv_init = openpyxl.worksheet.datavalidation.DataValidation.__init__
    def _patched_dv_init(self, *args, **kwargs):
        kwargs.pop('id', None)
        _orig_dv_init(self, *args, **kwargs)
    openpyxl.worksheet.datavalidation.DataValidation.__init__ = _patched_dv_init
except Exception:
    pass

# 导入参数配置加载器
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from parsers.param_config_loader import load_param_config, classify_param


@dataclass
class FieldParamMapping:
    """字段→参数映射"""
    dm_table: str
    dm_field: str
    report_table: str = ''
    report_field: str = ''
    param_code: str = ''
    param_category: str = ''  # business / personalization
    case_when_expr: str = ''
    ref_source: str = ''
    param_values: List[str] = field(default_factory=list)


def _join_toclob_fragments(content: str) -> str:
    """
    将 to_clob('...') || to_clob('...') 格式的PL/SQL拼接为完整SQL。
    
    ETL清洗代码中，长SQL被拆分为多个 to_clob('片段') || to_clob('片段') 的格式，
    每个片段内部用 '' 表示单引号转义。
    此函数提取所有片段内容并拼接，同时将 '' 还原为 '。
    """
    # 提取所有 to_clob('...') 片段内容
    # to_clob 内部用 '' 转义单引号
    toclob_pattern = re.compile(r"to_clob\('((?:[^']|'')*?)'\)", re.DOTALL)
    fragments = toclob_pattern.findall(content)
    
    if not fragments:
        # 无 to_clob 格式，返回原始内容
        return content
    
    # 拼接所有片段，将 '' 还原为 '
    joined = ''
    for frag in fragments:
        joined += frag.replace("''", "'")
    
    return joined


def _extract_dm_table_from_filename(fname: str) -> str:
    """
    从ETL文件名提取DM表名。
    
    支持格式：
    - DM_CTRC_XTCPLXRXXZBB_中信登指标清洗.sql → DM_CTRC_XTCPLXRXXZBB
    - URP_SRDT5_xxx_0_RPT.sql → URP_SRDT5_xxx
    """
    # 去掉 .sql 后缀
    name = fname.rsplit('.', 1)[0] if '.' in fname else fname
    
    # DM_ 前缀：取前三段作为表名（DM_CTRC_XTCPLXRXXZBB）
    if name.upper().startswith('DM_'):
        parts = name.split('_')
        # DM 表名通常是 DM_模块_表名，后面是中文描述
        # 需要判断哪些部分属于表名
        dm_parts = []
        for i, p in enumerate(parts):
            if i < 3:  # DM, CTRC, XTCPLXRXXZBB 等前3段
                dm_parts.append(p)
            elif p.isascii() and p.isupper():
                # 继续是英文大写，可能是表名的一部分
                dm_parts.append(p)
            else:
                # 中文或混合内容，视为描述部分
                break
        if dm_parts:
            return '_'.join(dm_parts)
    
    # URP_ 前缀类似处理
    if name.upper().startswith('URP_'):
        parts = name.split('_')
        urp_parts = []
        for i, p in enumerate(parts):
            if i < 3 or (p.isascii() and p.isupper()):
                urp_parts.append(p)
            else:
                break
        if urp_parts:
            return '_'.join(urp_parts)
    
    # 回退：取文件名（去后缀）
    return name


def parse_etl_cleansing_files(etl_dir: str, 
                                param_config: Dict) -> List[FieldParamMapping]:
    """
    扫描ETL清洗代码文件，提取 #[PARAM] 占位符与字段的关系。
    
    解析逻辑：
    1. 扫描每个 .sql 文件
    2. 从文件名提取 DM 表名
    3. 将 to_clob('...') || to_clob('...') 格式拼接为完整SQL
    4. 查找 #[PARAM] 占位符及其上下文
    5. 从 CASE WHEN #[PARAM] 结构提取参数控制逻辑
    6. 从 INSERT INTO 语句提取 DM 字段与 #[PARAM] 的关联
    
    Args:
        etl_dir: etl_code_list 目录路径
        param_config: 参数配置字典
        
    Returns:
        List[FieldParamMapping]
    """
    mappings = []
    
    sql_files = glob.glob(os.path.join(etl_dir, '*.sql'))
    if not sql_files:
        # Try uppercase
        sql_files = glob.glob(os.path.join(etl_dir, '*.SQL'))
    
    for fpath in sql_files:
        fname = os.path.basename(fpath)
        dm_table = _extract_dm_table_from_filename(fname)
        
        try:
            with open(fpath, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
        except Exception:
            continue
        
        # Step 1: 拼接 to_clob 片段为完整SQL
        joined_sql = _join_toclob_fragments(content)
        
        # Step 2: 提取 #[PARAM] 引用
        param_refs = re.findall(r"#\[([A-Z_0-9]+)\]", joined_sql, re.IGNORECASE)
        if not param_refs:
            # 继续用原始内容试一次
            param_refs = re.findall(r"#\[([A-Z_0-9]+)\]", content, re.IGNORECASE)
            if not param_refs:
                continue
            joined_sql = content  # 用原始内容
        
        unique_refs = list(set(param_refs))
        
        # Step 3: 在完整SQL中查找 CASE WHEN #[PARAM] 模式
        # 模式: CASE WHEN #[PARAM] = 'VALUE' THEN ... END [AS] alias
        case_pattern = re.compile(
            r"CASE\s+WHEN\s+['\"]?#\[(\w+)\]#?['\"]?\s*=\s*['\"]?(\w+)['\"]?\s+THEN\s+"
            r"(.+?)(?:\s+END\s+(?:AS\s+)?(\w+))",
            re.IGNORECASE | re.DOTALL
        )
        
        field_param_pairs = set()  # (dm_table, field_alias, param_code)
        
        for match in case_pattern.finditer(joined_sql):
            param_code = match.group(1).upper()
            param_value = match.group(2)
            then_expr = match.group(3)[:200]  # 截断
            field_alias = match.group(4) if match.group(4) else ''
            
            # 分类参数
            actual_code = param_code
            from parsers.param_config_loader import _get_equivalent_param
            equiv = _get_equivalent_param(param_code)
            if param_code not in param_config and equiv and equiv in param_config:
                actual_code = equiv
            
            if actual_code in param_config:
                category = param_config[actual_code].category
            else:
                category = classify_param(param_code, '', [])
            
            mapping = FieldParamMapping(
                dm_table=dm_table,
                dm_field=field_alias,
                param_code=param_code,
                param_category=category,
                case_when_expr=f"CASE WHEN #[{param_code}] = '{param_value}' THEN {then_expr[:100]}",
                ref_source=fname,
            )
            mappings.append(mapping)
            field_param_pairs.add((dm_table, field_alias, param_code))
        
        # Step 3b: 检测嵌套 CASE WHEN 中的参数引用
        # 当外层 CASE WHEN ... END AS field_alias 的 THEN 表达式中
        # 包含嵌套的 #[PARAM] 引用时，将嵌套参数也映射到同一 field_alias
        for match in case_pattern.finditer(joined_sql):
            outer_param_code = match.group(1).upper()
            field_alias = match.group(4) if match.group(4) else ''
            then_expr = match.group(3)
            
            if not field_alias or not then_expr:
                continue
            
            nested_params = re.findall(r"#\[([A-Z_0-9]+)\]", then_expr, re.IGNORECASE)
            for nested_code in nested_params:
                nested_code = nested_code.upper()
                if nested_code == outer_param_code:
                    continue
                
                key = (dm_table, field_alias, nested_code)
                if key in field_param_pairs:
                    continue
                
                actual_code = nested_code
                from parsers.param_config_loader import _get_equivalent_param
                equiv = _get_equivalent_param(nested_code)
                if nested_code not in param_config and equiv and equiv in param_config:
                    actual_code = equiv
                
                if actual_code in param_config:
                    category = param_config[actual_code].category
                else:
                    category = classify_param(nested_code, '', [])
                
                mapping = FieldParamMapping(
                    dm_table=dm_table,
                    dm_field=field_alias,
                    param_code=nested_code,
                    param_category=category,
                    case_when_expr=f"NESTED within CASE WHEN #[{outer_param_code}] ... END AS {field_alias}",
                    ref_source=fname,
                )
                mappings.append(mapping)
                field_param_pairs.add(key)
        
        # Step 4: 提取 INSERT INTO 字段列表，找到哪些字段被 #[PARAM] 控制
        # 从 INSERT INTO ... (FIELD1, FIELD2, ...) 中提取字段列表
        insert_pattern = re.compile(
            r"INSERT\s+INTO\s+\w+\.?(\w+)?\s*\(([^)]+)\)",
            re.IGNORECASE
        )
        insert_fields = []
        for match in insert_pattern.finditer(joined_sql):
            fields_str = match.group(2)
            insert_fields = [f.strip().rstrip(',') for f in fields_str.split(',') if f.strip()]
            break  # 只取第一个 INSERT INTO
        
        # Step 5: 如果没有找到 CASE WHEN 模式但文件包含 #[PARAM]，
        # 为每个 #[PARAM] 创建通用映射
        if not field_param_pairs:
            for param_code in unique_refs:
                actual_code = param_code
                from parsers.param_config_loader import _get_equivalent_param
                equiv = _get_equivalent_param(param_code)
                if param_code not in param_config and equiv and equiv in param_config:
                    actual_code = equiv
                
                if actual_code in param_config:
                    category = param_config[actual_code].category
                else:
                    category = classify_param(param_code, '', [])
                
                mapping = FieldParamMapping(
                    dm_table=dm_table,
                    dm_field='',  # 无法确定具体字段
                    param_code=param_code,
                    param_category=category,
                    case_when_expr=f'#[{param_code}] referenced in ETL file',
                    ref_source=fname,
                )
                mappings.append(mapping)
        
        # Step 6: 查找 JOIN 条件中的 #[PARAM] 引用
        join_param_pattern = re.compile(
            r"(?:AND|ON)\s+['\"]?#\[([A-Z_0-9]+)\]#?['\"]?\s*=\s*['\"]?([^'\"]+)['\"]?",
            re.IGNORECASE
        )
        for match in join_param_pattern.finditer(joined_sql):
            param_code = match.group(1).upper()
            param_value = match.group(2)
            
            key = (dm_table, '', param_code)
            if key in field_param_pairs:
                continue  # 已经记录过
            
            actual_code = param_code
            from parsers.param_config_loader import _get_equivalent_param
            equiv = _get_equivalent_param(param_code)
            if param_code not in param_config and equiv and equiv in param_config:
                actual_code = equiv
            
            if actual_code in param_config:
                category = param_config[actual_code].category
                values = param_config[actual_code].value_list
            else:
                category = classify_param(param_code, '', [])
                values = []
            
            mapping = FieldParamMapping(
                dm_table=dm_table,
                dm_field='',
                param_code=param_code,
                param_category=category,
                case_when_expr=f"JOIN condition: #[{param_code}] = '{param_value}'",
                ref_source=fname,
                param_values=values,
            )
            mappings.append(mapping)
    
    return mappings


def parse_inspection_sql_params(xlsx_path: str,
                                 param_config: Dict) -> List[FieldParamMapping]:
    """
    从巡检列SQL模板中提取 URP_PARAM_CONFIG 引用。
    
    Args:
        xlsx_path: my_poor_solution.xlsx 路径
        param_config: 参数配置字典
        
    Returns:
        List[FieldParamMapping]
    """
    mappings = []
    
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[22]:
            continue
        
        sql = str(row[22])
        if 'URP_PARAM_CONFIG' not in sql and '#[' not in sql:
            continue
        
        field_alias = str(row[1]) if row[1] else ''
        dm_table = str(row[12]) if row[12] else ''
        dm_field = str(row[13]) if row[13] else ''
        check_code = str(row[3]) if row[3] else ''
        
        # Extract URP_PARAM_CONFIG param_codes
        config_codes = re.findall(r"param_code\s*=\s*'([^']+)'", sql, re.IGNORECASE)
        hash_codes = re.findall(r"#\[([A-Z_0-9]+)\]", sql, re.IGNORECASE)
        all_codes = list(set(config_codes + hash_codes))
        
        for code in all_codes:
            # Classify
            actual_code = code
            from parsers.param_config_loader import _get_equivalent_param
            equiv = _get_equivalent_param(code)
            if code not in param_config and equiv and equiv in param_config:
                actual_code = equiv
            
            if actual_code in param_config:
                category = param_config[actual_code].category
                values = param_config[actual_code].value_list
            else:
                category = classify_param(code, '', [])
                values = []
            
            mapping = FieldParamMapping(
                dm_table=dm_table,
                dm_field=dm_field,
                report_table=field_alias.split('.')[0] if '.' in field_alias else '',
                report_field=field_alias.split('.')[1] if '.' in field_alias else '',
                param_code=code,
                param_category=category,
                case_when_expr='URP_PARAM_CONFIG scalar subquery' if code in config_codes else '#[PARAM] placeholder',
                ref_source=f'inspection:{check_code}',
                param_values=values,
            )
            mappings.append(mapping)
    
    wb.close()
    return mappings


def save_param_field_mapping(mappings: List[FieldParamMapping], output_path: str):
    """保存字段→参数映射到xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '字段参数映射'
    
    # Headers
    headers = ['DM_TABLE', 'DM_FIELD', 'REPORT_TABLE', 'REPORT_FIELD', 
               'PARAM_CODE', 'PARAM_CATEGORY', 'CASE_WHEN_EXPR', 
               'REF_SOURCE', 'PARAM_VALUES']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for i, m in enumerate(mappings, 2):
        ws.cell(row=i, column=1, value=m.dm_table)
        ws.cell(row=i, column=2, value=m.dm_field)
        ws.cell(row=i, column=3, value=m.report_table)
        ws.cell(row=i, column=4, value=m.report_field)
        ws.cell(row=i, column=5, value=m.param_code)
        ws.cell(row=i, column=6, value=m.param_category)
        ws.cell(row=i, column=7, value=m.case_when_expr)
        ws.cell(row=i, column=8, value=m.ref_source)
        ws.cell(row=i, column=9, value='|'.join(m.param_values))
    
    # Column widths
    col_widths = [25, 25, 25, 25, 25, 15, 60, 30, 30]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(65 + col - 1)].width = width
    
    wb.save(output_path)


def main():
    """主执行入口"""
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    
    # 路径配置
    etl_dir = os.path.join(base_dir, 'etl_code_list')
    param_config_path = os.path.join(base_dir, 'code_param_list.xlsx')
    inspection_path = os.path.join(base_dir, 'my_poor_solution.xlsx')
    output_path = os.path.join(base_dir, 'param_field_mapping.xlsx')
    
    print(f"=== ETL参数依赖提取 ===")
    print(f"ETL路径: {etl_dir}")
    print(f"参数配置: {param_config_path}")
    print(f"巡检列: {inspection_path}")
    print()
    
    # 1. 加载参数配置
    print("1. 加载参数配置...")
    param_config = load_param_config(param_config_path)
    print(f"   加载 {len(param_config)} 个参数")
    
    # 2. 扫描ETL清洗代码
    print("2. 扫描ETL清洗代码...")
    etl_mappings = parse_etl_cleansing_files(etl_dir, param_config)
    print(f"   找到 {len(etl_mappings)} 条字段→参数映射")
    
    # 3. 扫描巡检列SQL模板
    print("3. 扫描巡检列SQL模板...")
    inspection_mappings = parse_inspection_sql_params(inspection_path, param_config)
    print(f"   找到 {len(inspection_mappings)} 条参数引用")
    
    # 4. 合并去重
    all_mappings = etl_mappings + inspection_mappings
    
    # 按字段+参数代码去重
    seen = set()
    unique_mappings = []
    for m in all_mappings:
        key = (m.dm_table, m.dm_field, m.param_code)
        if key not in seen:
            seen.add(key)
            unique_mappings.append(m)
    
    print(f"4. 去重后: {len(unique_mappings)} 条唯一映射")
    
    # 5. 统计
    business_count = sum(1 for m in unique_mappings if m.param_category == 'business')
    personal_count = sum(1 for m in unique_mappings if m.param_category == 'personalization')
    
    # 唯一参数代码
    unique_params = set(m.param_code for m in unique_mappings)
    unique_dm_fields = set((m.dm_table, m.dm_field) for m in unique_mappings if m.dm_field)
    
    print(f"\n=== 统计 ===")
    print(f"唯一映射数: {len(unique_mappings)}")
    print(f"业务参数映射: {business_count}")
    print(f"个性化参数映射: {personal_count}")
    print(f"涉及参数代码: {len(unique_params)} 个")
    print(f"涉及DM字段: {len(unique_dm_fields)} 个")
    
    # 6. 保存
    save_param_field_mapping(unique_mappings, output_path)
    print(f"\n5. 参数映射已保存: {output_path}")
    
    # 7. 展开预估
    # 找出涉及业务参数的DM字段，计算每个字段的展开数
    business_field_params = {}
    for m in unique_mappings:
        if m.param_category == 'business' and m.dm_field:
            key = (m.dm_table, m.dm_field)
            if key not in business_field_params:
                business_field_params[key] = set()
            business_field_params[key].add(m.param_code)
    
    total_expansion = 0
    for key, params in business_field_params.items():
        expansion = 1
        for p in params:
            actual_p = p
            from parsers.param_config_loader import _get_equivalent_param
            equiv = _get_equivalent_param(p)
            if p not in param_config and equiv and equiv in param_config:
                actual_p = equiv
            if actual_p in param_config:
                expansion *= max(len(param_config[actual_p].param_values), 2)
            else:
                expansion *= 2
        total_expansion += expansion
    
    print(f"\n=== 展开预估 ===")
    print(f"涉及业务参数的DM字段: {len(business_field_params)} 个")
    print(f"ETL层面展开预估: +{total_expansion} SQL")


if __name__ == '__main__':
    main()