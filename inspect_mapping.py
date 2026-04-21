"""临时脚本：检查 my_poor_solution.xlsx 的巡检列sheet结构"""
import openpyxl
import os

# Patch DataValidation
_orig_dv_init = openpyxl.worksheet.datavalidation.DataValidation.__init__
def _patched_dv_init(self, *args, **kwargs):
    kwargs.pop('id', None)
    _orig_dv_init(self, *args, **kwargs)
openpyxl.worksheet.datavalidation.DataValidation.__init__ = _patched_dv_init

xlsx_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'my_poor_solution.xlsx')
print(f"Loading: {xlsx_path}")
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
print(f"Sheet names: {wb.sheetnames}")
print()

for sn in wb.sheetnames:
    if '巡检' in sn or '检列' in sn:
        ws = wb[sn]
        print(f"=== Sheet: {sn} ===")
        print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
        # Print header
        header = []
        for cell in ws[1]:
            header.append(str(cell.value) if cell.value else '')
        print(f"  Header ({len(header)} cols): {header[:20]}")
        # Print first 5 data rows
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True)):
            print(f"  Row {i+2}: {list(row)[:20]}")
        # Print total data rows
        data_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                data_count += 1
        print(f"  Total data rows: {data_count}")
        print()

# Also check the DM字段映射 sheet (index 2 or 3)
for sn in wb.sheetnames:
    if 'DM' in sn or '映射' in sn:
        ws = wb[sn]
        print(f"=== Sheet: {sn} ===")
        print(f"  Rows: {ws.max_row}, Cols: {ws.max_column}")
        header = []
        for cell in ws[1]:
            header.append(str(cell.value) if cell.value else '')
        print(f"  Header ({len(header)} cols): {header[:15]}")
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=5, values_only=True)):
            print(f"  Row {i+2}: {list(row)[:15]}")
        print()

wb.close()