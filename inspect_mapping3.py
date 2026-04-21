"""统计巡检列sheet中的关键字段映射覆盖率"""
import openpyxl
import os

_orig_dv_init = openpyxl.worksheet.datavalidation.DataValidation.__init__
def _patched_dv_init(self, *args, **kwargs):
    kwargs.pop('id', None)
    _orig_dv_init(self, *args, **kwargs)
openpyxl.worksheet.datavalidation.DataValidation.__init__ = _patched_dv_init

xlsx_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'my_poor_solution.xlsx')
wb = openpyxl.load_workbook(xlsx_path, data_only=True)

# 巡检列 sheet - 第一列是字段序号，col[1]是字段别名（如 URP_TPRT_CPJBXX.XTDJXTCPBM）
# col[12]是DM表，col[13]是DM字段，col[17]是STG表中文名，col[18]是STG字段中文名
# col[19]是取数来源(即STG表名)，col[20]是取数来源字段
# col[21]是关联体系ID（如 T2_TCMP_SUBPROJECTINFO.DM_CTRC_CPJBXXZBB）
# col[22]是关联内容（即完整SQL模板）

# 读取report_check_list.xlsx的报表名
check_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'report_check_list.xlsx')
wb2 = openpyxl.load_workbook(check_path, read_only=True)
check_tables = set()
for ws in wb2.worksheets:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) > 1 and row[1]:
            check_tables.add(str(row[1]).strip())
wb2.close()

print(f"report_check_list 中的报表名数量: {len(check_tables)}")
print(f"前10个报表名: {sorted(list(check_tables))[:10]}")

# 从巡检列 sheet 读取
ws = wb[wb.sheetnames[0]]
field_alias_set = set()  # 字段别名集合
stg_dm_pairs = {}  # 关联体系ID → 映射信息
dm_stg_map = {}  # DM表 → [{stg_table, stg_key, sql_template, data_source}]
direct_match = {}  # 字段别名 → {dm_table, dm_field, stg_table, stg_field, sql_template}

row_count = 0
has_sql_template = 0
has_stg_table = 0
has_dm_table = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or row[0] is None:
        continue
    row_count += 1
    
    field_alias = str(row[1]).strip() if row[1] else ""  # URP_TPRT_CPJBXX.XTDJXTCPBM
    check_code_coord = str(row[3]).strip() if row[3] else ""  # 如 DQ01.A0001
    dm_table = str(row[12]).strip() if row[12] else ""  # DM表名
    dm_field = str(row[13]).strip() if row[13] else ""  # DM字段名
    stg_table_cn = str(row[17]).strip() if row[17] else ""  # STG表中文名
    stg_field_cn = str(row[18]).strip() if row[18] else ""  # STG字段中文名
    stg_table = str(row[19]).strip() if row[19] else ""  # 取数来源（STG表名）
    stg_field = str(row[20]).strip() if row[20] else ""  # 取数来源字段
    link_id = str(row[21]).strip() if row[21] else ""  # 关联体系ID
    sql_template = str(row[22]).strip() if row[22] else ""  # 关联内容（SQL模板）
    
    if field_alias:
        field_alias_set.add(field_alias)
    
    if dm_table:
        has_dm_table += 1
    if stg_table and stg_table != '0':
        has_stg_table += 1
    if sql_template and sql_template != '0' and sql_template != 'None':
        has_sql_template += 1
    
    # 构建映射
    if link_id and stg_table and stg_table != '0':
        if link_id not in stg_dm_pairs:
            stg_dm_pairs[link_id] = {
                'link_id': link_id,
                'stg_table': stg_table,
                'dm_table': dm_table,
                'stg_field': stg_field,
                'dm_field': dm_field,
                'sql_template': sql_template,
                'stg_table_cn': stg_table_cn,
                'count': 0
            }
        stg_dm_pairs[link_id]['count'] += 1
    
    # 直接映射：字段别名 → 详细信息
    if field_alias and dm_table:
        if field_alias not in direct_match:
            direct_match[field_alias] = {
                'dm_table': dm_table,
                'dm_field': dm_field,
                'stg_table': stg_table if stg_table != '0' else '',
                'stg_field': stg_field,
                'stg_table_cn': stg_table_cn,
                'stg_field_cn': stg_field_cn if row[18] else '',
                'link_id': link_id,
                'sql_template': sql_template if sql_template != '0' and sql_template != 'None' else '',
                'field_alias': field_alias,
                'check_code': check_code_coord,
            }

print(f"\n巡检列 总数据行: {row_count}")
print(f"  有DM表的行: {has_dm_table}")
print(f"  有STG表的行: {has_stg_table}")
print(f"  有SQL模板的行: {has_sql_template}")
print(f"  唯一字段别名: {len(field_alias_set)}")
print(f"  关联体系ID数: {len(stg_dm_pairs)}")
print(f"  直接映射条目: {len(direct_match)}")

# 检查report_check_list中的报表名能否匹配到巡检列
matched = 0
unmatched_tables = []
for table in sorted(list(check_tables))[:30]:
    # 找在巡检列中是否有匹配
    found = False
    for alias, info in direct_match.items():
        if table in alias:
            found = True
            break
    if found:
        matched += 1
    else:
        unmatched_tables.append(table)

print(f"\n前30个报表名匹配测试: 匹配={matched}, 不匹配={len(unmatched_tables)}")
if unmatched_tables[:10]:
    print(f"不匹配的报表名前10: {unmatched_tables[:10]}")

# 看看report_check_list中的report_table和巡检列中DATA_TABLE_NAME的匹配关系
# 从巡检列前条件 sheet 检查
ws3 = wb[wb.sheetnames[2]]
print(f"\n\n巡检列前条件 sheet:")
data_tables = set()
for row in ws3.iter_rows(min_row=2, values_only=True):
    if row and row[2]:
        dt = str(row[2]).strip()
        data_tables.add(dt)

print(f"  唯一DATA_TABLE_NAME数: {len(data_tables)}")
# 与report_check_list的报表名交集
overlap = check_tables & data_tables
print(f"  与report_check_list交集: {len(overlap)}")
print(f"  报表名仅在report_check_list中: {len(check_tables - data_tables)}")
print(f"  报表名仅在巡检列前条件中: {len(data_tables - check_tables)}")

# 检查字段别名格式的报表关系
# 字段别名格式: URP_TPRT_CPJBXX.XTDJXTCPBM → 报表名=URP_TPRT_CPJBXX, 字段名=XTDJXTCPBM
report_from_alias = set()
for alias in field_alias_set:
    if '.' in alias:
        parts = alias.split('.')
        report_from_alias.add(parts[0])

print(f"\n字段别名中提取的报表名数量: {len(report_from_alias)}")
overlap2 = check_tables & report_from_alias
print(f"  与report_check_list交集: {len(overlap2)}")
print(f"  report_check_list总报表数: {len(check_tables)}")

# 示例匹配
print("\n=== 匹配示例 ===")
sample_checks = ['URP_SRDT5_GDXXB', 'URP_TPRT_CPJBXX', 'URP_TPRT_CPTZ', 'DM_CTRC_CPJBXXZBB']
for table in sample_checks:
    matches = [(alias, info) for alias, info in direct_match.items() if table in alias]
    print(f"\n  报表名 '{table}' 匹配到 {len(matches)} 条:")
    for alias, info in matches[:3]:
        print(f"    {alias} → DM={info['dm_table']}, STG={info['stg_table']}, field={info['stg_field']}")

wb.close()