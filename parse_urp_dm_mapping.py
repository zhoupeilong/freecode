"""
parse_urp_dm_mapping.py - 从 URP ETL 脚本解析 URP报表字段↔DM指标表字段的映射关系

核心逻辑：
1. 扫描 urp_code_list 目录下的 SQL 文件
2. 解析 INSERT INTO ... SELECT 语句中的字段别名
3. 提取 URP表名、URP字段名、DM表名、DM字段名的映射
4. 输出 xlsx 文件

映射提取规则：
- INSERT INTO ${V_USER_XTRPT}.URP_xxxx (col_list) → URP表名 = URP_xxxx
- FROM 子句中的 DM 表名 → DM表名 (如 DM_CTRC_CPJBXXZBB)
- SELECT 中的列别名 → URP字段名 = 别名, DM字段名 = 源表达式中的字段名

SQL 模式识别：
- 模式1: T.XTDJXTCPBM XTDJXTCPBM  → DM=XTDJXTCPBM, URP=XTDJXTCPBM (同名)
- 模式2: t.shareholder_type gdlx    → DM=SHAREHOLDER_TYPE, URP=GDLX (异名)
- 模式3: expression AS alias         → DM=expression中的字段, URP=alias
- 模式4: V_COMPANY_NAME xtjgmc      → DM=系统变量(无映射), URP=XTJGMC
- 模式5: CASE WHEN ... END AS alias  → DM=CASE中的字段, URP=alias
"""

import os
import re
import openpyxl
from typing import List, Dict, Tuple, Optional

# 目标URP表清单
TARGET_URP_TABLES = [
    "URP_SRDT5_1104TJZBDZB", "URP_SRDT5_ALLSUBJECTS", "URP_SRDT5_CPJBXXB",
    "URP_SRDT5_CPQSXXB", "URP_SRDT5_CPTZB", "URP_SRDT5_DKGLXXB",
    "URP_SRDT5_FDCGMB", "URP_SRDT5_FXXMXXB", "URP_SRDT5_FXYSZBDZB",
    "URP_SRDT5_GDXXB", "URP_SRDT5_GLFXXB", "URP_SRDT5_GQJHHQYFELGLXXB",
    "URP_SRDT5_GYDBHTB", "URP_SRDT5_GYDBRXXB", "URP_SRDT5_GYDZYWXXB",
    "URP_SRDT5_GYJYDSJGB", "URP_SRDT5_GYJYDSJRCPB", "URP_SRDT5_GYJYDSZZRB",
    "URP_SRDT5_GYNBKMDZB", "URP_SRDT5_GYYDSGLGXB", "URP_SRDT5_GYYYHTXXB",
    "URP_SRDT5_GYYYJYLSB", "URP_SRDT5_GYZHXXB", "URP_SRDT5_HLWDKHZJGXXB",
    "URP_SRDT5_HLWDKJKRMXB", "URP_SRDT5_JGXXB", "URP_SRDT5_JYDSGLGXB",
    "URP_SRDT5_JYDSJGB", "URP_SRDT5_JYDSJRCPB", "URP_SRDT5_JYDSZRRB",
    "URP_SRDT5_QTFBZHZQGLXXB", "URP_SRDT5_QYSZGMB", "URP_SRDT5_TDMDZTGLXXB",
    "URP_SRDT5_TDZCSSYQGLXXB", "URP_SRDT5_TSLYGMB", "URP_SRDT5_WTRJQCCXXB",
    "URP_SRDT5_XMRZDYGXB", "URP_SRDT5_XMRZXXB", "URP_SRDT5_XSFSXXB",
    "URP_SRDT5_XTCPLXRXXB", "URP_SRDT5_XTCPXSMXB", "URP_SRDT5_XTCPZCXXB",
    "URP_SRDT5_XTCPZHXXB", "URP_SRDT5_XTCPZZKJQKMB", "URP_SRDT5_XTCYRHZXXB",
    "URP_SRDT5_XTDBHTXXB", "URP_SRDT5_XTDBRXXB", "URP_SRDT5_XTDZYWXXB",
    "URP_SRDT5_XTFYXXB", "URP_SRDT5_XTGMDGXMFQCSJB", "URP_SRDT5_XTGMDGXMZLSJB",
    "URP_SRDT5_XTKHJGB", "URP_SRDT5_XTKHJRCPB", "URP_SRDT5_XTKHZRRB",
    "URP_SRDT5_XTNBKMDZB", "URP_SRDT5_XTSYQXXB", "URP_SRDT5_XTSYQZRXXB",
    "URP_SRDT5_XTYYJYLSB", "URP_SRDT5_YGXXB", "URP_SRDT5_YWFLXXB",
    "URP_SRDT5_ZCZQHGLXXB", "URP_SRDT5_ZJLYLBXXB", "URP_TPRT_CPJBXX",
    "URP_TPRT_CPTZ", "URP_TPRT_CZCZQYXX", "URP_TPRT_DBHTXX", "URP_TPRT_DBRXX",
    "URP_TPRT_DCZCCTXX", "URP_TPRT_DCZCXX", "URP_TPRT_DZYWXX", "URP_TPRT_FDCGM",
    "URP_TPRT_FXXMXX", "URP_TPRT_GLJYXX", "URP_TPRT_GTSTRXX",
    "URP_TPRT_GYCSXTDCZCXX", "URP_TPRT_GYCSXTZCXX", "URP_TPRT_HLWDKXX",
    "URP_TPRT_JCRXX", "URP_TPRT_JYDSXX", "URP_TPRT_LXRXX", "URP_TPRT_QYSZGM",
    "URP_TPRT_SYQXX", "URP_TPRT_SYSYRXX", "URP_TPRT_TSLYGM", "URP_TPRT_TZGWXX",
    "URP_TPRT_WTRJQCCXX", "URP_TPRT_XMRZDYGXB", "URP_TPRT_XMRZXX",
    "URP_TPRT_XSFSXX", "URP_TPRT_XTFYXX", "URP_TPRT_XTGMDGXMFQCSJ",
    "URP_TPRT_XTGMDGXMZLSJ", "URP_TPRT_YDTJXX", "URP_TPRT_YWFLXX",
    "URP_TPRT_YZZXRXX", "URP_TPRT_ZCZQHXX", "URP_TPRT_ZHXX",
    "URP_TPRT_ZJLYLBXX",
]

# 需要排除的 URP 系统字段（非业务字段）
URP_SYSTEM_COLUMNS = {
    'ID', 'REPORT_FLOW_ID', 'BUSINESS_REPORT_CODE', 'DATA_BELONG',
    'REPORT_DATE', 'SEND_STATUS', 'FEDD_BACK_REMARK', 'INDEX_NUM',
    'ROW_NUM', 'CJRQ', 'DATA_PARTITION', 'BBZ',
}

# 也排除小写的系统字段
URP_SYSTEM_COLUMNS_LOWER = {c.lower() for c in URP_SYSTEM_COLUMNS}


def extract_sql_content(file_path: str) -> str:
    """
    从 PL/SQL 包装中提取纯 SQL 内容。
    
    处理两种格式：
    1. to_clob('...') || to_clob('...') 串联格式
    2. 纯 SQL 格式
    """
    with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
        raw = f.read()
    
    # 提取所有 to_clob('...') 内容
    clob_pattern = re.compile(r"to_clob\('((?:[^']|'')*?)'\)", re.DOTALL)
    clob_matches = clob_pattern.findall(raw)
    
    if clob_matches:
        # 拼接 clob 内容（去除 Oracle 转义 '' → '）
        sql_parts = []
        for part in clob_matches:
            # Oracle '' 转义为单引号
            unescaped = part.replace("''", "'")
            sql_parts.append(unescaped)
        sql_content = '\n'.join(sql_parts)
    else:
        # 没有 to_clob，直接使用原始内容
        sql_content = raw
    
    # 去除 SQL 行注释（-- 开头的注释行），避免中文乱码污染表达式
    sql_content = _remove_sql_comments(sql_content)
    
    return sql_content


def _remove_sql_comments(sql: str) -> str:
    """
    去除 SQL 中的行注释（-- 开头的注释），避免中文乱码污染。
    
    保留注释中的有用信息（如字段中文名），但只保留英文和合法字符。
    对于含乱码的注释行，直接去除整行。
    """
    lines = sql.split('\n')
    cleaned = []
    for line in lines:
        stripped = line.lstrip()
        # 如果整行是注释
        if stripped.startswith('--'):
            # 检查注释是否包含乱码（替换字符 U+FFFD 或 GBK 乱码特征）
            # 乱码行通常包含大量 \ufffd 或连续的非 ASCII 非 CJK 字符
            has_garbled = False
            garbled_count = 0
            for ch in stripped:
                if ch == '\ufffd':
                    has_garbled = True
                    break
                # 检测 GBK 乱码：Latin-1 范围内的非 ASCII 字符（0x80-0xFF）
                # 在 UTF-8 中这些字节如果被错误解码会变成孤立的高位字符
                if 0x80 <= ord(ch) <= 0xFF:
                    garbled_count += 1
            # 如果注释整行都是乱码或有超过2个乱码字符，去除
            if has_garbled or garbled_count > 2:
                continue
            # 保留不含乱码的注释行（如纯英文注释）
            cleaned.append(line)
            continue
        
        # 对于代码行中的行尾注释，去除注释部分但保留代码
        # 查找行中的 -- 注释（但不包括字符串常量内的 --）
        comment_pos = _find_line_comment_pos(line)
        if comment_pos >= 0:
            # 检查注释部分是否有乱码
            comment_part = line[comment_pos:]
            has_garbled = '\ufffd' in comment_part
            garbled_count = sum(1 for ch in comment_part if 0x80 <= ord(ch) <= 0xFF)
            if has_garbled or garbled_count > 2:
                # 去除有乱码的注释，保留代码部分
                cleaned.append(line[:comment_pos].rstrip())
                continue
        
        cleaned.append(line)
    
    return '\n'.join(cleaned)


def _find_line_comment_pos(line: str) -> int:
    """
    查找 SQL 行中 -- 注释的位置，但不考虑字符串常量内的 --。
    返回 -1 表示没有行注释。
    """
    in_string = False
    i = 0
    while i < len(line):
        ch = line[i]
        if ch == "'" and not in_string:
            in_string = True
        elif ch == "'" and in_string:
            # 检查是否是转义的 '' （Oracle 风格）
            if i + 1 < len(line) and line[i + 1] == "'":
                i += 1  # 跳过下一个引号
            else:
                in_string = False
        elif ch == '-' and not in_string and i + 1 < len(line) and line[i + 1] == '-':
            return i
        i += 1
    return -1


def _clean_expression(expr: str) -> str:
    """
    清理 SQL 表达式中的注释和多余空白，确保输出干净无乱码。
    
    1. 去除行内 -- 注释
    2. 去除乱码字符 (U+FFFD 和 GBK 乱码)
    3. 压缩多余空白
    4. 移除前导的换行和空格
    """
    if not expr:
        return expr
    
    # 去除行内注释
    comment_pos = _find_line_comment_pos(expr)
    if comment_pos >= 0:
        expr = expr[:comment_pos].rstrip()
    
    # 去除乱码字符 (U+FFFD 替换字符)
    expr = expr.replace('\ufffd', '')
    
    # 去除 GBK 乱码：在 UTF-8 中错误解码产生的连续高位字符
    # 这些通常是 0x80-0xFF 范围内的孤立字节，看起来像 "快速" "产品" 等乱码
    # 但我们需要保留合法的 CJK 字符 (U+4E00-U+9FFF) 和正常 ASCII
    cleaned = []
    for ch in expr:
        cp = ord(ch)
        # 保留：ASCII、CJK统一汉字、常见符号
        if cp <= 0x7F:  # ASCII
            cleaned.append(ch)
        elif 0x4E00 <= cp <= 0x9FFF:  # CJK基本汉字
            cleaned.append(ch)
        elif 0x3000 <= cp <= 0x303F:  # CJK标点
            cleaned.append(ch)
        elif 0xFF00 <= cp <= 0xFFEF:  # 全角字符
            cleaned.append(ch)
        elif 0x0080 <= cp <= 0x00FF:  # Latin-1 补充（GBK乱码区域）
            # 跳过这些乱码字符
            continue
        else:
            # 其他 Unicode 字符（如带音调的拉丁字母等）保留
            cleaned.append(ch)
    
    expr = ''.join(cleaned)
    
    # 压缩连续空白为单个空格
    expr = re.sub(r'\s+', ' ', expr)
    
    return expr.strip()


def parse_insert_select(sql_content: str) -> List[Dict]:
    """
    从 SQL 内容中解析 INSERT INTO ... SELECT 语句的字段映射。
    
    返回列表：每项包含 {
        'urp_table': str,     # URP报表表名
        'dm_table': str,      # DM指标表名
        'urp_column': str,    # URP报表字段名
        'dm_column': str,     # DM指标表字段名（可能为空表达式）
        'dm_expression': str,  # DM字段表达式（完整）
        'is_system_column': bool,  # 是否系统字段
    }
    """
    mappings = []
    
    # 查找所有 INSERT INTO 语句
    # 模式1: INSERT INTO ${V_USER_XTRPT}.URP_xxxx (col_list) SELECT ...
    # 模式2: INSERT INTO ${V_USER_XTRPT}.URP_xxxx\n  (col_list)\n  SELECT ...
    
    insert_pattern = re.compile(
        r'INSERT\s+INTO\s+(?:\$\{V_USER_XTRPT\}\.)?(URP_\w+)\s*\(([^)]+)\)\s*SELECT\s+(.+?)(?:FROM\s+.+|$)',
        re.IGNORECASE | re.DOTALL
    )
    
    # 用于 SRDT5 类型：INSERT ... SELECT ... FROM ... 模式
    # 先找出所有 INSERT INTO URP_xxxx 片段
    # 更灵活的匹配
    insert_re = re.compile(
        r'INSERT\s+INTO\s+(?:\$\{V_USER_XTRPT\}\.)?(\w+)\s*\(([^)]+)\)',
        re.IGNORECASE | re.DOTALL
    )
    
    for m in insert_re.finditer(sql_content):
        urp_table_raw = m.group(1).upper()
        col_list_str = m.group(2)
        
        # 只处理 URP 表
        if not urp_table_raw.startswith('URP_'):
            continue
        
        # 解析列名列表
        urp_columns = [c.strip().upper() for c in col_list_str.split(',') if c.strip()]
        # 过滤空值
        urp_columns = [c for c in urp_columns if c and not c.startswith('--')]
        
        if not urp_columns:
            continue
        
        # 找到对应的 SELECT 部分
        # 从 INSERT INTO ... (col_list) 之后开始，找到 SELECT ... FROM ...
        select_start = m.end()
        # 跳过 SELECT 前的空白
        remaining = sql_content[select_start:].lstrip()
        
        if not remaining.upper().startswith('SELECT') and not remaining.upper().startswith('\nSELECT'):
            # 可能有空格或换行
            sel_match = re.search(r'\bSELECT\b', remaining, re.IGNORECASE)
            if not sel_match:
                continue
            remaining = remaining[sel_match.start():]
        
        # 找到 SELECT ... FROM 之间的内容
        # 需要处理子查询、CASE WHEN 等嵌套结构
        select_body = _extract_select_body(remaining)
        
        if not select_body:
            continue
        
        # 从 SELECT body 中提取 DM 表名和字段映射
        dm_table = _extract_dm_table(remaining)
        
        # 按列位置提取映射
        select_items = _parse_select_items(select_body)
        
        # 匹配列位置
        for i, col in enumerate(urp_columns):
            if i < len(select_items):
                expr, alias = select_items[i]
                # 确定UPR字段名
                urp_col = col if col else (alias if alias else '')
                
                # 跳过系统字段
                is_system = urp_col.upper() in URP_SYSTEM_COLUMNS or urp_col.lower() in URP_SYSTEM_COLUMNS_LOWER
                
                # 从表达式中提取 DM 字段名
                dm_field = _extract_dm_field_from_expr(expr)
                
                # 清理表达式中的行内注释（保留纯代码）
                clean_expr = _clean_expression(expr.strip())
                
                if urp_col and urp_col not in URP_SYSTEM_COLUMNS:
                    mappings.append({
                        'urp_table': urp_table_raw,
                        'urp_column': urp_col,
                        'dm_table': dm_table or '',
                        'dm_column': dm_field or '',
                        'dm_expression': clean_expr,
                        'is_system_column': is_system,
                    })
    
    return mappings


def _extract_select_body(sql_after_select: str) -> Optional[str]:
    """提取 SELECT 到 FROM 之间的内容（处理嵌套子查询）"""
    if not sql_after_select.upper().lstrip().startswith('SELECT'):
        return None
    
    # 找到顶层 FROM 关键字（不在子查询中的）
    depth = 0
    from_pos = -1
    i = 0
    upper = sql_after_select.upper()
    
    while i < len(sql_after_select):
        if upper[i:i+6] == 'SELECT' and (i == 0 or not upper[i-1].isalnum()):
            depth += 1
            i += 6
        elif upper[i:i+4] == 'FROM' and (i == 0 or not upper[i-1].isalnum()) and depth == 1:
            from_pos = i
            break
        elif sql_after_select[i] == '(':
            depth += 1
            i += 1
        elif sql_after_select[i] == ')':
            depth -= 1
            i += 1
        else:
            i += 1
    
    if from_pos > 0:
        return sql_after_select[6:from_pos].strip()  # 跳过 SELECT 本身
    
    # 如果没有找到顶层 FROM，返回空
    return None


def _extract_dm_table(sql_text: str) -> Optional[str]:
    """从 SQL 文本中提取 DM 表名"""
    # 模式: FROM ${V_USER_XTDM}.DM_xxx 或 FROM ${V_USER_XTDW}.DW_xxx
    from_pattern = re.compile(
        r'FROM\s+(?:\$\{V_USER_\w+\}\.)?(DM_\w+|DW_\w+)',
        re.IGNORECASE
    )
    m = from_pattern.search(sql_text)
    if m:
        return m.group(1).upper()
    
    # 模式: from (subquery) T - 需要找子查询中的 DM 表
    # 模式: FROM ${V_USER_XTDM}.DM_CTRC_xxx t1
    from_pattern2 = re.compile(
        r'FROM\s+(?:\$\{V_USER_\w+\}\.)?(\bDM_\w+)',
        re.IGNORECASE
    )
    m = from_pattern2.search(sql_text)
    if m:
        return m.group(1).upper()
    
    return None


def _parse_select_items(select_body: str) -> List[Tuple[str, Optional[str]]]:
    """
    解析 SELECT 列表中的每一项。
    
    返回 [(expression, alias), ...]
    
    处理：
    - T.XTDJXTCPBM XTDJXTCPBM
    - t.shareholder_type gdlx
    - expression AS alias
    - V_FINL_LICS_CODE || ... AS id
    - CASE WHEN ... END AS alias
    """
    items = []
    
    # 按逗号分割，但需要处理 CASE WHEN ... END 中的逗号和函数中的逗号
    # 使用括号深度和 CASE 深度来分割
    parts = _split_select_items(select_body)
    
    alias_pattern = re.compile(
        r'\bAS\s+(\w+)\s*$',
        re.IGNORECASE
    )
    
    # 简单别名模式: expression alias (无 AS)
    simple_alias = re.compile(
        r'^(\S+(?:\.\S+)?)\s+(\w+)\s*$'
    )
    
    # 函数调用别名模式: func(...) alias
    func_alias = re.compile(
        r'^(\w+\([^)]*\))\s+(\w+)\s*$',
        re.IGNORECASE
    )
    
    for part in parts:
        part = part.strip()
        if not part:
            continue
        
        # 去除行内注释（如 "t.xtdjxtcpbm xtdjxtcpbm -- 信托产品编码"）
        comment_pos = _find_line_comment_pos(part)
        if comment_pos >= 0:
            # 只保留注释前的代码部分
            code_part = part[:comment_pos].rstrip()
            # 如果整行都是注释（除空白外），跳过
            if not code_part:
                continue
            part = code_part
        
        # 尝试 AS alias 模式
        m = alias_pattern.search(part)
        if m:
            alias = m.group(1)
            expression = part[:m.start()].strip()
            items.append((expression, alias))
            continue
        
        # 尝试简单别名模式 expression alias
        m = simple_alias.match(part)
        if m:
            expr = m.group(1)
            alias = m.group(2)
            # 确认不是关键字
            if alias.upper() not in ('FROM', 'WHERE', 'AND', 'OR', 'JOIN', 'ON', 'INNER', 'LEFT', 'RIGHT', 'CASE', 'WHEN', 'THEN', 'ELSE', 'END', 'ORDER', 'GROUP', 'HAVING', 'INTO', 'SET', 'VALUES', 'SELECT', 'AS'):
                items.append((expr, alias))
                continue
        
        # 没有别名
        items.append((part, None))
    
    return items


def _split_select_items(select_body: str) -> List[str]:
    """
    按顶层逗号分割 SELECT 列表项，正确处理括号嵌套和 CASE...END 嵌套。
    """
    items = []
    current = []
    depth = 0       # 括号深度
    case_depth = 0  # CASE 嵌套深度
    upper = select_body.upper()
    i = 0
    
    while i < len(select_body):
        ch = select_body[i]
        
        # 处理字符串常量（跳过）
        if ch == "'":
            # 找到配对的引号
            j = i + 1
            while j < len(select_body):
                if select_body[j] == "'" and j + 1 < len(select_body) and select_body[j+1] == "'":
                    j += 2  # Oracle 转义 ''
                    continue
                elif select_body[j] == "'":
                    j += 1
                    break
                else:
                    j += 1
            current.append(select_body[i:j])
            i = j
            continue
        
        if ch == '(':
            depth += 1
            current.append(ch)
            i += 1
        elif ch == ')':
            depth -= 1
            current.append(ch)
            i += 1
        elif ch == ',' and depth == 0 and case_depth == 0:
            items.append(''.join(current))
            current = []
            i += 1
        else:
            # 检查 CASE/END 关键字
            if upper[i:i+4] == 'CASE' and (i == 0 or not upper[i-1].isalnum()):
                case_depth += 1
                current.append(ch)
                i += 1
            elif upper[i:i+3] == 'END' and (i + 3 >= len(select_body) or not upper[i+3].isalnum()):
                case_depth -= 1
                current.append(ch)
                i += 1
            else:
                current.append(ch)
                i += 1
    
    if current:
        items.append(''.join(current))
    
    return items


def _extract_dm_field_from_expr(expression: str) -> str:
    """
    从 SQL 表达式中提取 DM 字段名。
    
    处理模式：
    - T.XTDJXTCPBM → XTDJXTCPBM
    - t.shareholder_type → SHAREHOLDER_TYPE  
    - ${V_USER_XTDW}.fn_translatedict('CTRC_0044',t.glfs) → GLFS
    - CASE WHEN ... THEN T.JGGS END → JGGS
    - V_COMPANY_NAME → (系统变量，无DM字段)
    - ''#[DW_ZXDJGID]'' → (参数变量，无DM字段)
    - 表达式 → 提取第一个 T.xxx 或 t.xxx 模式的字段
    """
    expr = expression.strip()
    
    # 移除 SQL 函数包装，提取内部 T.xxx 字段
    # 优先提取最内层的 T.xxx 或 t.xxx 模式
    t_field_pattern = re.compile(r'[Tt]\.(\w+)', re.IGNORECASE)
    matches = t_field_pattern.findall(expr)
    
    if matches:
        # 返回第一个匹配的 T.xxx 中的字段名
        return matches[0].upper()
    
    # 尝试匹配 T1.xxx 模式 (子查询别名)
    t1_field_pattern = re.compile(r'[Tt]1\.(\w+)', re.IGNORECASE)
    matches = t1_field_pattern.findall(expr)
    if matches:
        return matches[0].upper()
    
    # 如果是纯列名 (无表别名前缀)
    # 检查是否是简单的列名 (全大写字母+数字+下划线)
    simple_col = re.compile(r'^[A-Z][A-Z0-9_]*$', re.IGNORECASE)
    if simple_col.match(expr):
        return expr.upper()
    
    # 系统变量或常量
    if '${' in expr or "'#[" in expr or expr.startswith("'") or expr.startswith('V_'):
        return ''
    
    # 复杂表达式，尝试从 AS 之前的表达式中提取
    return ''


def parse_urp_file(file_path: str) -> List[Dict]:
    """
    解析单个 URP SQL 文件，提取字段映射。
    
    Returns: 映射列表
    """
    sql_content = extract_sql_content(file_path)
    
    # 从文件名提取 URP 表名
    filename = os.path.basename(file_path)
    # 文件名格式: URP_TPRT_CPJBXX_1_中信登指标.sql 或 URP_SRDT5_CPJBXXB_0_RPT.sql
    # 提取 URP_xxxx 部分
    name_match = re.match(r'(URP_\w+?)_\d+_', filename)
    urp_table_from_file = name_match.group(1).upper() if name_match else ''
    
    mappings = parse_insert_select(sql_content)
    
    # 如果解析不到 URP 表名，使用文件名中的
    if mappings:
        # 验证解析到的 URP 表名与文件名一致
        for m in mappings:
            if not m['urp_table'] and urp_table_from_file:
                m['urp_table'] = urp_table_from_file
    elif urp_table_from_file:
        # 更激进的解析：直接搜索 SQL 中的 INSERT INTO URP_xxx
        mappings = _fallback_parse(sql_content, urp_table_from_file)
    
    return mappings


def _fallback_parse(sql_content: str, urp_table: str) -> List[Dict]:
    """
    后备解析：直接搜索 INSERT INTO URP_xxx (col_list) 模式，
    然后匹配 SELECT 中的列别名。
    """
    mappings = []
    
    # 查找 INSERT INTO 语句
    insert_pattern = re.compile(
        r'INSERT\s+INTO\s+(?:\$\{V_USER_\w+\}\.)?(' + urp_table + r')\s*\(([^)]+)\)',
        re.IGNORECASE | re.DOTALL
    )
    
    m = insert_pattern.search(sql_content)
    if not m:
        return []
    
    col_list_str = m.group(2)
    urp_columns = [c.strip().upper() for c in col_list_str.split(',') if c.strip()]
    urp_columns = [c for c in urp_columns if c and not c.startswith('--')]
    
    # 提取 DM 表名
    dm_table = _extract_dm_table(sql_content)
    
    # 查找 SELECT 列表
    select_start = m.end()
    remaining = sql_content[select_start:].lstrip()
    
    # 找 SELECT
    sel_match = re.search(r'\bSELECT\b', remaining, re.IGNORECASE)
    if not sel_match:
        return []
    
    select_text = remaining[sel_match.start():]
    select_body = _extract_select_body(select_text)
    if not select_body:
        return []
    
    select_items = _parse_select_items(select_body)
    
    for i, col in enumerate(urp_columns):
        is_system = col in URP_SYSTEM_COLUMNS
        dm_field = ''
        dm_expr = ''
        
        if i < len(select_items):
            expr, alias = select_items[i]
            dm_field = _extract_dm_field_from_expr(expr)
            dm_expr = _clean_expression(expr.strip())
        
        if not is_system:
            mappings.append({
                'urp_table': urp_table,
                'urp_column': col,
                'dm_table': dm_table or '',
                'dm_column': dm_field,
                'dm_expression': dm_expr,
                'is_system_column': False,
            })
    
    return mappings


def main():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    urp_dir = os.path.join(base_dir, 'urp_code_list')
    output_path = os.path.join(base_dir, 'urp_dm_field_mapping.xlsx')
    
    print(f"扫描目录: {urp_dir}")
    
    # 收集所有映射
    all_mappings = {}  # key: (urp_table, urp_column) → mapping dict
    
    # 按目标 URP 表处理
    for urp_table in TARGET_URP_TABLES:
        # 查找匹配的 SQL 文件
        # 优先使用 _0_RPT 或 _1_RPT 文件（主逻辑）
        matching_files = []
        for fname in os.listdir(urp_dir):
            if fname.startswith(urp_table + '_') and fname.endswith('.sql'):
                matching_files.append(os.path.join(urp_dir, fname))
        
        if not matching_files:
            continue
        
        # 优先选择 RPT 文件或中信登指标文件
        priority_file = None
        for f in matching_files:
            basename = os.path.basename(f)
            if '_0_RPT' in basename or '_1_RPT' in basename or '_1_中信登指标' in basename:
                priority_file = f
                break
        
        if not priority_file:
            # 退而求其次，选第一个
            priority_file = matching_files[0]
        
        # 解析文件
        mappings = parse_urp_file(priority_file)
        
        for m in mappings:
            key = (m['urp_table'], m['urp_column'])
            if key not in all_mappings:
                all_mappings[key] = m
            else:
                # 如果有 DM 字段映射而之前没有，更新
                existing = all_mappings[key]
                if m['dm_column'] and not existing['dm_column']:
                    all_mappings[key] = m
    
    # 生成输出
    print(f"\n解析结果: {len(all_mappings)} 条映射关系")
    
    # 按 URP 表分组统计
    table_counts = {}
    for key, m in all_mappings.items():
        urp_tbl = m['urp_table']
        table_counts[urp_tbl] = table_counts.get(urp_tbl, 0) + 1
    
    print("\n各表映射数量:")
    for tbl in sorted(table_counts.keys()):
        print(f"  {tbl}: {table_counts[tbl]} 条")
    
    # 写入 xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "URP_DM字段映射"
    
    # 表头
    headers = ['ID', 'URP_TABLE_NAME', 'URP_COLUMN_NAME', 'DM_TABLE_NAME', 'DM_COLUMN_NAME', 'DM_EXPRESSION', '备注']
    ws.append(headers)
    
    # 数据行
    for key in sorted(all_mappings.keys()):
        m = all_mappings[key]
        urp_table = m['urp_table']
        urp_column = m['urp_column']
        dm_table = m['dm_table']
        dm_column = m['dm_column']
        dm_expr = m['dm_expression']
        
        # ID = URP_TABLE_NAME.URP_COLUMN_NAME
        row_id = f"{urp_table}.{urp_column}"
        
        # 备注
        note = ''
        if not dm_column:
            note = 'DM字段为系统变量或常量表达式'
        elif dm_column == urp_column:
            note = '同名映射'
        else:
            note = '异名映射'
        
        ws.append([row_id, urp_table, urp_column, dm_table, dm_column, dm_expr, note])
    
    # 格式化列宽
    col_widths = [50, 30, 30, 30, 30, 60, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # 设置表头样式
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    wb.save(output_path)
    print(f"\n输出文件: {output_path}")


if __name__ == "__main__":
    main()