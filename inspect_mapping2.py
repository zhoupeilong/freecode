"""深入检查 巡检列 sheet的完整列结构"""
import openpyxl
import os

_orig_dv_init = openpyxl.worksheet.datavalidation.DataValidation.__init__
def _patched_dv_init(self, *args, **kwargs):
    kwargs.pop('id', None)
    _orig_dv_init(self, *args, **kwargs)
openpyxl.worksheet.datavalidation.DataValidation.__init__ = _patched_dv_init

xlsx_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'my_poor_solution.xlsx')
wb = openpyxl.load_workbook(xlsx_path, data_only=True)

# 巡检列 sheet - 完整header和前20行
ws = wb[wb.sheetnames[0]]  # 巡检列
print(f"=== Sheet: {wb.sheetnames[0]} ===")
print(f"Total rows: {ws.max_row}, Total cols: {ws.max_column}")

# Print full header
header = []
for cell in ws[1]:
    header.append(str(cell.value) if cell.value else '')
print(f"\nFull header ({len(header)} cols):")
for i, h in enumerate(header):
    print(f"  [{i}] {h}")

# Print first 10 data rows - ALL columns
print("\n=== First 10 data rows (all columns) ===")
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True)):
    print(f"\n--- Row {i+2} ---")
    for j, val in enumerate(row):
        if val is not None:
            val_str = str(val)
            if len(val_str) > 120:
                val_str = val_str[:120] + "..."
            print(f"  [{j}] {header[j] if j < len(header) else '?'}: {val_str}")

# Check the DM字段映射 sheet
ws2 = wb[wb.sheetnames[3]]  # 关联到DM字段映射
print(f"\n\n=== Sheet: {wb.sheetnames[3]} ===")
print(f"Total rows: {ws2.max_row}, Total cols: {ws2.max_column}")
header2 = []
for cell in ws2[1]:
    header2.append(str(cell.value) if cell.value else '')
print(f"\nFull header ({len(header2)} cols):")
for i, h in enumerate(header2):
    print(f"  [{i}] {h}")

# Print first 5 data rows
print("\n=== First 5 data rows ===")
for i, row in enumerate(ws2.iter_rows(min_row=2, max_row=6, values_only=True)):
    print(f"\n--- Row {i+2} ---")
    for j, val in enumerate(row):
        if val is not None:
            val_str = str(val)
            if len(val_str) > 120:
                val_str = val_str[:120] + "..."
            print(f"  [{j}] {header2[j] if j < len(header2) else '?'}: {val_str}")

# Check sheet names[2] - 巡检列前条件
ws3 = wb[wb.sheetnames[2]]
print(f"\n\n=== Sheet: {wb.sheetnames[2]} ===")
print(f"Total rows: {ws3.max_row}, Total cols: {ws3.max_column}")
header3 = []
for cell in ws3[1]:
    header3.append(str(cell.value) if cell.value else '')
print(f"\nFull header ({len(header3)} cols):")
for i, h in enumerate(header3[:20]):
    print(f"  [{i}] {h}")

# Print first 5 data rows
print("\n=== First 5 data rows ===")
for i, row in enumerate(ws3.iter_rows(min_row=2, max_row=6, values_only=True)):
    print(f"\n--- Row {i+2} ---")
    for j, val in enumerate(row):
        if val is not None:
            val_str = str(val)
            if len(val_str) > 120:
                val_str = val_str[:120] + "..."
            print(f"  [{j}] {header3[j] if j < len(header3) else '?'}: {val_str}")

wb.close()