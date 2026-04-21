"""
generate_stg_checks.py - STG 巡检脚本生成主入口（v2 + v0.7参数展开）

核心变更：
- 使用巡检列 sheet 作为主映射源（字段别名精确匹配）
- 优先使用巡检列中的完整 SQL 模板
- 仅在无 SQL 模板时才拼装 SQL
- v0.7: 参数控制字段展开 — 业务参数按值组合生成多个SQL，个性化参数保留URP_PARAM_CONFIG

工作流程：
1. 读取 report_check_list.xlsx → ReportCheckInfo 列表
2. 从 my_poor_solution.xlsx 加载巡检列映射（主映射源）
2d. 从 code_param_list.xlsx 加载参数配置（v0.7新增）
3. 读取 stg_key_list.xlsx 获取业务主键
4. 对每条检查项：
   a. 通过字段别名(report_table.report_field)查找巡检列映射
   b. 如有 SQL 模板 → 直接使用模板 + 变量替换
   c. 如无 SQL 模板 → 基于 InspectionRow 信息拼装
   d. 如有业务参数 → 展开为多份SQL（v0.7新增）
5. 批量写入 SUCxxxx.sql 文件
6. 记录生成日志
"""

import os
import sys
import re
import json
from datetime import datetime
from typing import Dict, List, Optional

# 添加项目路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from parsers.report_check_parser import load_report_checks, ReportCheckInfo
from parsers.etl_mapper import (
    load_my_poor_solution, ETLMapping, InspectionRow, STGDWMapping,
    find_mapping_for_check, get_stg_dw_mapping
)
from parsers.condition_builder import build_extra_condition
from parsers.param_config_loader import (
    load_param_config, classify_params_in_sql, expand_business_params,
    get_all_param_codes, build_param_value_label, ParamDefinition,
    _get_equivalent_param
)
from utils.logger import Logger
from utils.file_writer import BatchWriter
import openpyxl


# ============================================================
# 常量定义
# ============================================================

# 固定业务辅助字段（项目相关）
BUSINESS_FIELDS = """
        proj.count_proj_code,
        proj.count_proj_name,
        urp3_xtdw.fn_company_dict(proj.trust_manager, '1') trust_manager,
        urp3_xtdw.fn_company_dict(proj.deal_manage, '1') deal_manage,
        urp3_xtdw.fn_company_dict(proj.trust_manager_b, '1') trust_manager_b,
        urp3_xtdw.fn_company_dict(proj.trust_manager_c, '1') trust_manager_c,
        urp3_xtdw.fn_company_dict(proj.beit_dept, '2') beit_dept,"""

# 无项目关联报表的业务辅助字段
NO_PROJECT_FIELDS = """
        NULL count_proj_code,
        NULL count_proj_name,
        NULL trust_manager,
        NULL deal_manage,
        NULL trust_manager_b,
        NULL trust_manager_c,
        NULL beit_dept,"""

# 报送标识字段
REPORT_FLAG_FIELDS = """
        k.report_flag_tprt,
        k.report_flag_east5"""

def load_stg_key_list(xlsx_path: str) -> dict:
    """读取 stg_key_list.xlsx，返回 STG表名 → STG_KEY表达式 映射"""
    import openpyxl
    if not os.path.exists(xlsx_path):
        return {}
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            result[str(row[0]).strip()] = str(row[1]).strip()
    wb.close()
    return result


def load_stg_table_info(table_xlsx_path: str) -> dict:
    """读取 table_structure_list.xlsx，返回 STG表.字段 → 中文名"""
    import openpyxl
    if not os.path.exists(table_xlsx_path):
        return {}
    wb = openpyxl.load_workbook(table_xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 4 and row[1] and row[2]:
            table_name = str(row[1]).strip()
            col_name = str(row[2]).strip()
            col_cn = str(row[3]).strip() if row[3] else ""
            upper_table = table_name.upper()
            if upper_table.startswith('T2_') or upper_table.startswith('TS_'):
                if table_name not in result:
                    result[table_name] = {}
                result[table_name][col_name] = col_cn
    wb.close()
    return result


def is_project_related(check_info: ReportCheckInfo) -> bool:
    """判断检查项是否与项目相关"""
    non_project_keywords = ['YG', '员工', 'GLF', '关联方', 'GLJY', '关联交易', '313']
    check_name = check_info.check_name or ""
    report_table = check_info.report_table or ""
    for kw in non_project_keywords:
        if kw in check_name.upper() or kw in report_table.upper():
            return False
    return True


def is_project_related_by_name(report_table: str) -> bool:
    """根据报表名判断是否与项目相关"""
    # 非项目报表模式：员工、关联方等
    non_project_patterns = [
        'YGXXB', 'GLF', 'GLJY', 'YG_', 'YGB',
        '_YG', 'GLFXXB', 'YGXX',
    ]
    upper = report_table.upper()
    for p in non_project_patterns:
        if p in upper:
            return False
    return True


def _fill_sql_template(template: str, check_info: ReportCheckInfo,
                       insp_row: InspectionRow, stg_key: str,
                       stg_table_info: dict, has_project: bool) -> str:
    """
    使用巡检列中的 SQL 模板，通过变量替换生成完整 SQL。

    模板中的占位符：
    - ${CHECK_NAME} → 勾稽名称
    - ${STG_TABLE_NAME} → STG表名
    - ${STG_TABLE_NAME_CN} → STG表中文名
    - ${STG_COL_NAME} → STG字段名
    - ${STG_COL_NAME_CN} → STG字段中文名
    - ${STG_KEY} → STG业务主键表达式
    - ${K_EXPR} → 检验表达式（如 trim(k.XTDJXTCPBM) is null）
    """
    sql = template

    # 获取STG表中文名
    stg_cn = stg_table_info.get(insp_row.stg_table, {})
    stg_table_cn = insp_row.stg_table_cn or insp_row.stg_table
    stg_field_cn = insp_row.stg_field_cn or insp_row.stg_field

    # 替换占位符
    sql = sql.replace('${CHECK_NAME}', check_info.check_name)
    sql = sql.replace('${STG_TABLE_NAME}', insp_row.stg_table)
    sql = sql.replace('${STG_TABLE_NAME_CN}', stg_table_cn)
    sql = sql.replace('${STG_COL_NAME}', insp_row.stg_field or check_info.report_field)
    sql = sql.replace('${STG_COL_NAME_CN}', stg_field_cn)

    # STG_KEY替换
    if stg_key:
        sql = sql.replace('${STG_KEY}', stg_key)
    else:
        sql = sql.replace('${STG_KEY}', "'业务主键[待补充]'")

    # K_EXPR替换：检验表达式
    check_expr = insp_row.check_expression
    if not check_expr:
        # 默认：trim(k.字段) is null
        dm_field = insp_row.dm_field or check_info.report_field
        if check_info.is_conditional:
            # 条件必填
            pre_cond = _build_precondition(check_info, insp_row)
            if pre_cond:
                check_expr = f"{pre_cond}\n  AND trim(k.{dm_field}) IS NULL"
            else:
                check_expr = f"trim(k.{dm_field}) IS NULL"
        else:
            check_expr = f"trim(k.{dm_field}) IS NULL"

    sql = sql.replace('${K_EXPR}', check_expr)

    # ===========================
    # 借用参考行SQL模板时的字段替换
    # ===========================
    # 当SQL模板来自同DM表的参考行时，模板中硬编码了参考行的字段值。
    # 我们需要谨慎地只替换"检核相关的字段"，而不是JOIN条件或通用字段。
    #
    # 安全替换策略：
    # 1. WHERE子句中的 "trim(k.XXX) is null" → 替换为当前DM字段
    # 2. SELECT子句中的 "XXX stg_col_value" → 替换为当前STG字段
    # 3. 注释头部中添加当前检查项的正确check_name
    # 不替换（保持模板原始值）：
    # - JOIN条件中的 t.字段 或 k.字段（如 k.count_proj_code, t.project_code）
    # - stg_table_name, stg_col_name（除非使用${占位符}）
    # - stg_col_name_cn, check_name 等字符串常量

    cur_dm_field = insp_row.dm_field or check_info.report_field

    # 1. 替换WHERE子句中硬编码的检核条件
    if insp_row.dm_field:
        cur_expr = f"trim(k.{cur_dm_field}) IS NULL"

        # 查找 "trim(k.XXX) is null" 模式并替换非当前DM字段的条件
        k_field_pattern = re.compile(r'k\.(\w+)', re.IGNORECASE)
        template_fields = set(k_field_pattern.findall(sql))
        
        protected_fields = {'CAL_DATE', 'COUNT_PROJ_CODE', 'REPORT_FLAG_TPRT', 
                          'REPORT_FLAG_EAST5', 'BUSI_SCOPE', 'PROJECT_CODE', 
                          'PROJ_CODE', 'DATA_SOURCE'}
        for field_in_template in template_fields:
            field_upper = field_in_template.upper()
            if field_upper != cur_dm_field.upper() and field_upper not in protected_fields:
                # 只替换 trim(k.XXX) is null 这种检核条件模式
                ref_pattern = re.compile(
                    rf'trim\s*\(\s*k\s*\.\s*{re.escape(field_in_template)}\s*\)\s+[Ii][Ss]\s+[Nn][Uu][Ll][Ll]',
                    re.IGNORECASE
                )
                sql = ref_pattern.sub(cur_expr, sql)

    # 2. 注释头部（保证报表字段、DM字段、检查编号、勾稽名称完整）
    header = _build_sql_header(check_info, insp_row)
    if not sql.startswith('--'):
        # 没有注释头部，直接添加
        sql = header + sql

    return sql


def _build_precondition(check_info: ReportCheckInfo, insp_row: InspectionRow) -> str:
    """为条件必填字段构建前置条件"""
    parts = []

    if insp_row.pre_table and insp_row.pre_table != '0':
        # 使用巡检列中的前置条件信息
        pre_table = insp_row.pre_table
        pre_col = insp_row.pre_column
        pre_val = insp_row.pre_value

        if pre_col and pre_val:
            # 简化：前置条件已包含在SQL模板的K_EXPR中或巡检列的检验表达式中
            return ""  # 由K_EXPR处理

    if check_info.precondition:
        return f"-- 条件: {check_info.precondition} (需要人工确认)"

    return ""


def generate_sql_from_template(check_info: ReportCheckInfo,
                                insp_row: InspectionRow,
                                stg_key: str,
                                stg_table_info: dict,
                                logger=None) -> str:
    """
    使用巡检列中的 SQL 模板生成巡检 SQL。

    判断逻辑：
    - 如果 insp_row 是精确匹配（字段别名完全对应），直接使用SQL模板
    - 如果 insp_row 是DM字段映射补充且有借用模板，使用模板+字段替换
      （保留参考行的JOIN条件、data_source过滤等结构）
    - 如果 insp_row 是DM字段映射补充且无模板，使用通用拼装

    Args:
        check_info: 检查项信息
        insp_row: 巡检列匹配行
        stg_key: STG主键表达式
        stg_table_info: STG表字段中文名映射
        logger: 日志记录器

    Returns:
        生成的 SQL 字符串
    """
    is_exact_match = not insp_row.is_from_dm_supplement
    has_project = is_project_related(check_info)

    # 精确匹配 → 使用SQL模板（模板中的字段值是正确的）
    if is_exact_match and insp_row.sql_template:
        sql = _fill_sql_template(insp_row.sql_template, check_info, insp_row,
                                  stg_key, stg_table_info, has_project)

        # 非项目相关报表需要替换项目辅助字段
        if not has_project:
            sql = _replace_project_fields(sql)

        return sql

    # DM补充行 → 如有借用模板，使用模板+字段替换？否则通用拼装
    if not is_exact_match and insp_row.sql_template:
        sql = _fill_sql_template(insp_row.sql_template, check_info, insp_row,
                                  stg_key, stg_table_info, has_project)

        # 替换 SELECT 子句中借用模板的硬编码字段值
        cur_stg_field = insp_row.stg_field or check_info.report_field
        stg_cn = stg_table_info.get(insp_row.stg_table, {})
        stg_field_cn = insp_row.stg_field_cn or stg_cn.get(cur_stg_field, cur_stg_field)
        sql = _replace_stg_field_in_select(sql, cur_stg_field,
                                            stg_field_cn=stg_field_cn,
                                            check_name=check_info.check_name)

        # 添加DM补充标记注释（供人工审核）
        sql = _add_dm_supplement_comment(sql, insp_row)

        if not has_project:
            sql = _replace_project_fields(sql)

        return sql

    # 无模板 → 使用通用拼装
    return _assemble_sql(check_info, insp_row, stg_key, stg_table_info, has_project, logger)


def _replace_stg_field_in_select(sql: str, cur_stg_field: str,
                                   stg_field_cn: str = '',
                                   check_name: str = '') -> str:
    """
    替换 SELECT 子句中借用模板的硬编码字段值（DM补充行专用）。

    当DM补充行借用参考行的SQL模板时，模板中的以下字段是硬编码的：
    - stg_col_value: t.REF_FIELD stg_col_value → t.CUR_FIELD stg_col_value
    - stg_col_name: 'REF_FIELD' stg_col_name → 'CUR_FIELD' stg_col_name
    - stg_col_name_cn: '参考中文名' stg_col_name_cn → '当前中文名' stg_col_name_cn
    - check_name: '参考名称' check_name → '当前名称' check_name

    Args:
        sql: SQL字符串
        cur_stg_field: 当前检查项的STG字段名
        stg_field_cn: 当前STG字段中文名（可选）
        check_name: 当前勾稽名称（可选）

    Returns:
        替换后的SQL字符串
    """
    # 1. 替换 stg_col_value: (t.REF_FIELD 或 REF_FIELD) stg_col_value → cur stg_col_value
    sql = re.sub(
        r'(?:\w+\.)?\w+\s+stg_col_value\b',
        f't.{cur_stg_field} stg_col_value',
        sql
    )

    # 2. 替换 stg_col_name: 'REF_FIELD' stg_col_name → 'CUR_FIELD' stg_col_name
    if cur_stg_field:
        sql = re.sub(
            r"'[^']*'\s+stg_col_name\b",
            f"'{cur_stg_field}' stg_col_name",
            sql
        )

    # 3. 替换 stg_col_name_cn: '参考中文名' stg_col_name_cn → '当前中文名' stg_col_name_cn
    if stg_field_cn:
        sql = re.sub(
            r"'[^']*'\s+stg_col_name_cn\b",
            f"'{stg_field_cn}' stg_col_name_cn",
            sql
        )

    # 4. 替换 check_name: '参考名称' check_name → '当前名称' check_name
    #    仅替换SELECT子句中的字符串字面量，不影响注释行
    if check_name:
        sql = re.sub(
            r"'[^']*'\s+check_name\b",
            f"'{check_name}' check_name",
            sql
        )

    return sql


def _add_dm_supplement_comment(sql: str, insp_row: InspectionRow) -> str:
    """
    为DM补充行生成的SQL添加警告注释标记。

    标记内容：
    - 提醒人工审核JOIN条件和关联逻辑
    - 说明借用模板的来源DM表
    """
    comment = (
        f"-- ⚠️ DM补充: 借用同DM表({insp_row.dm_table})的SQL模板，"
        f"已替换检核字段为({insp_row.dm_field})\n"
        f"-- 请人工确认: JOIN条件、data_source过滤、STG字段({insp_row.stg_field})是否正确\n"
    )
    # 在"报表字段"注释行之后插入DM补充标记
    if '--报表字段：' in sql:
        # 找到"-- 检查编号:"行位置，在其前插入
        lines = sql.split('\n')
        insert_pos = None
        for i, line in enumerate(lines):
            if line.startswith('-- 检查编号:'):
                insert_pos = i
                break
        if insert_pos is not None:
            lines.insert(insert_pos, comment.rstrip())
            return '\n'.join(lines)
    # 回退：在第二行插入注释
    if sql.startswith('--'):
        lines = sql.split('\n', 1)
        if len(lines) > 1:
            return lines[0] + '\n' + comment + lines[1]
        return lines[0] + '\n' + comment
    return comment + sql


def load_field_param_mapping(xlsx_path: str, param_config: Dict[str, ParamDefinition],
                             logger=None) -> Dict[tuple, List[str]]:
    """
    从 param_field_mapping.xlsx 加载字段级参数映射，构建 (dm_table, dm_field) → [param_code, ...] 查找表。
    
    只返回 business 类型的参数代码，且仅返回可展开的参数（在 param_config 中有值列表的）。
    """
    if not os.path.exists(xlsx_path):
        if logger:
            logger.warn(f"字段参数映射文件不存在: {xlsx_path}")
        return {}
    
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    
    col_dm_table = headers.index('DM_TABLE') if 'DM_TABLE' in headers else None
    col_dm_field = headers.index('DM_FIELD') if 'DM_FIELD' in headers else None
    col_param_code = headers.index('PARAM_CODE') if 'PARAM_CODE' in headers else None
    col_category = headers.index('PARAM_CATEGORY') if 'PARAM_CATEGORY' in headers else None
    
    if any(v is None for v in [col_dm_table, col_dm_field, col_param_code, col_category]):
        if logger:
            logger.warn(f"字段参数映射文件列头不匹配: {headers}")
        return {}
    
    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        dm_table = row[col_dm_table] or ''
        dm_field = row[col_dm_field] or ''
        param_code = row[col_param_code] or ''
        category = row[col_category] or ''
        
        if not dm_table or not dm_field or not param_code:
            continue
        if category != 'business':
            continue
        
        actual_code = param_code
        equiv = _get_equivalent_param(param_code)
        if param_code not in param_config and equiv and equiv in param_config:
            actual_code = equiv
        
        if actual_code not in param_config or not param_config[actual_code].param_values:
            continue
        
        key = (dm_table.strip().upper(), dm_field.strip().upper())
        if key not in lookup:
            lookup[key] = []
        if param_code not in lookup[key]:
            lookup[key].append(param_code)
    
    wb.close()
    
    if logger:
        logger.info(f"  字段参数映射: {len(lookup)} 个字段有业务参数, "
                     f"涉及 {sum(len(v) for v in lookup.values())} 个参数映射")
    
    return lookup


def load_etl_lineage_config(json_path: str, logger=None) -> dict:
    """
    加载ETL血缘映射配置（etl_lineage_config.json）。
    
    返回结构:
    {
      "dm_field_lineage": {
        "DM_TABLE.DM_FIELD": {
          "params": {
            "PARAM_CODE": {
              "controls": "join_path" | "field_mapping",
              "values": { "VALUE": { ... } }
            }
          }
        }
      },
      "dw_stg_mapping": {
        "DW_TABLE": { "DW_FIELD": {"stg_table":..., "stg_field":..., "stg_field_cn":...} }
      }
    }
    """
    if not os.path.exists(json_path):
        if logger:
            logger.warn(f"ETL血缘配置文件不存在: {json_path}")
        return {"dm_field_lineage": {}, "dw_stg_mapping": {}}
    
    with open(json_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    lineage_count = len(config.get('dm_field_lineage', {}))
    dw_count = len(config.get('dw_stg_mapping', {}))
    
    if logger:
        logger.info(f"  ETL血缘配置: {lineage_count} 个DM字段有血缘映射, {dw_count} 个DW表有STG映射")
    
    return config


def _build_stg_null_check(stg_field_expr: str) -> str:
    """
    根据STG字段表达式生成逻辑自洽的空值判断条件。
    
    简单字段: trim(t.EXT_FIELD_1) IS NULL
    NVL组合:   NVL(trim(t.EXT_FIELD_1), trim(t.USER_CODE)) IS NULL
    
    遵循 §9.8: STG字段表达式与空值判断必须逻辑自洽。
    """
    # 检测 NVL(expr1, expr2) 模式
    nvl_match = re.match(r'^NVL\((\w+\.\w+),\s*(\w+\.\w+)\)$', stg_field_expr.strip(), re.IGNORECASE)
    if nvl_match:
        field_a = nvl_match.group(1)
        field_b = nvl_match.group(2)
        return f"NVL(trim({field_a}), trim({field_b})) IS NULL"
    else:
        # 简单字段: t.EXT_FIELD_1
        return f"trim({stg_field_expr.strip()}) IS NULL"


def _build_stg_col_name_cn(stg_field_expr: str, lineage_value: dict) -> str:
    """
    根据STG字段表达式确定 stg_col_name_cn。
    
    简单字段: 直接用 lineage 中的 stg_col_name_cn
    NVL组合:   取优先字段（NVL第一个参数）的中文名 + "(优先)"
    """
    nvl_match = re.match(r'^NVL\((\w+\.\w+),\s*(\w+\.\w+)\)$', stg_field_expr.strip(), re.IGNORECASE)
    if nvl_match:
        # NVL组合 - 使用 stg_col_name_cn（配置中已指定为优先字段的中文名）
        return lineage_value.get('stg_col_name_cn', '')
    else:
        return lineage_value.get('stg_col_name_cn', '')


def rebuild_sql_with_lineage(
    base_sql: str,
    check_info: ReportCheckInfo,
    insp_row: InspectionRow,
    param_combo: Dict[str, str],
    param_config: Dict[str, ParamDefinition],
    lineage_config: dict,
    stg_key: str,
    stg_table_info: dict,
    has_project: bool,
    logger=None
) -> Optional[str]:
    """
    根据血缘配置和参数组合，重建完整的巡检SQL。
    
    当血缘配置中存在该DM字段的映射时，根据参数组合重建：
    1. JOIN链（受 join_path 类型参数控制）
    2. STG表/字段（受 field_mapping 类型参数控制）
    3. WHERE中的参数过滤条件
    4. 空值判断条件（与STG字段表达式逻辑自洽）
    
    返回 None 表示该字段无血缘配置，应使用原始SQL。
    """
    dm_table = (insp_row.dm_table or '').strip().upper()
    dm_field = (check_info.report_field or '').strip().upper()
    lineage_key = f"{dm_table}.{dm_field}"
    
    field_lineage = lineage_config.get('dm_field_lineage', {}).get(lineage_key)
    if not field_lineage:
        return None  # 无血缘配置，不重建
    
    params = field_lineage.get('params', {})
    
    # 收集每个维度（join_path / field_mapping）的当前值配置
    join_config = None
    field_config = None
    
    for param_code, param_def in params.items():
        actual_code = param_code
        equiv = _get_equivalent_param(param_code)
        if param_code not in param_config and equiv and equiv in param_config:
            actual_code = equiv
        
        combo_value = param_combo.get(param_code, param_combo.get(actual_code, ''))
        
        if not combo_value:
            if logger:
                logger.warn(f"  血缘重建: 参数 {param_code} 在组合中未找到值，跳过")
            continue
        
        values = param_def.get('values', {})
        value_config = values.get(str(combo_value))
        if not value_config:
            if logger:
                logger.warn(f"  血缘重建: 参数 {param_code}={combo_value} 在血缘配置中未找到")
            continue
        
        controls = param_def.get('controls', '')
        if controls == 'join_path':
            join_config = value_config
        elif controls == 'field_mapping':
            field_config = value_config
    
    if not field_config:
        return None  # 没有字段映射配置，无法重建
    
    # 构建重建SQL
    stg_table = field_config.get('stg_table', '')
    stg_col_name = field_config.get('stg_col_name', '')
    stg_col_name_cn = field_config.get('stg_col_name_cn', '')
    stg_field_expr = field_config.get('stg_field_expr', '')
    
    # stg_key: 根据STG表确定
    stg_key_expr = stg_key  # 保留原始stg_key逻辑
    
    # JOIN条件
    join_sql = ''
    if join_config:
        join_sql = join_config.get('join_condition_sql', '')
        # 替换ETL中的别名 T → proj, T2/T4 → emp
        join_sql = join_sql.replace('T.', 'proj.')
    
    # 空值判断条件
    null_check = _build_stg_null_check(stg_field_expr)
    
    # stg_col_name_cn for NVL
    stg_col_name_cn = _build_stg_col_name_cn(stg_field_expr, field_config)
    
    # WHERE参数过滤条件
    param_filters = []
    for param_code in params.keys():
        actual_code = param_code
        equiv = _get_equivalent_param(param_code)
        if param_code not in param_config and equiv and equiv in param_config:
            actual_code = equiv
        combo_value = param_combo.get(param_code, param_combo.get(actual_code, ''))
        if combo_value:
            label = build_param_value_label(param_code, combo_value, param_config)
            param_filters.append(
                f"   AND (SELECT param_value FROM urp3.URP_PARAM_CONFIG t \n"
                f"         WHERE param_code = '{param_code}' AND status = '1') = '{combo_value}'"
                f"  --{label}"
            )
    
    # 组装SQL
    check_name = check_info.check_name
    
    # 判断是否项目相关
    if has_project:
        business_fields = BUSINESS_FIELDS
    else:
        business_fields = NO_PROJECT_FIELDS
    
    # 确定STG表的stg_key
    # T2_TCMP_EMPLOYEEINFO → '用户编码['||user_code||'],用户名称['||user_name||']'
    stg_key_for_table = stg_key_expr
    if not stg_key_for_table and stg_table == 'T2_TCMP_EMPLOYEEINFO':
        stg_key_for_table = "'用户编码['||user_code||'],用户名称['||user_name||']'"
    
    sql_parts = []
    sql_parts.append(f"select distinct  --勾稽名称")
    sql_parts.append(f"       '{check_name}' check_name,")
    sql_parts.append(f"       --STG要素")
    sql_parts.append(f"       '{stg_table}' stg_table_name,")
    sql_parts.append(f"        '{_get_stg_table_cn(stg_table)}' stg_table_name_cn,")
    sql_parts.append(f"        '{stg_col_name}' stg_col_name,")
    sql_parts.append(f"        '{stg_col_name_cn}' stg_col_name_cn,    ")
    sql_parts.append(f"       {stg_field_expr} stg_col_value,")
    sql_parts.append(f"       {stg_key_for_table} stg_key,")
    sql_parts.append(f"       --固定字段")
    sql_parts.append(business_fields)
    sql_parts.append(REPORT_FLAG_FIELDS)
    
    # FROM + JOIN
    sql_parts.append(f" from  urp3_tusp.{dm_table} k")
    sql_parts.append("inner join urp3_xtdw.dw_d_count_proj_info proj")
    sql_parts.append("ON     k.count_proj_code = proj.count_proj_code  AND proj.data_source='TCMP'")
    
    if join_sql:
        sql_parts.append(join_sql)
    
    sql_parts.append(f"inner join urp3_xtstg.{stg_table} t")
    sql_parts.append(f"\ton emp.employee_id=t.user_code ")
    
    # WHERE
    sql_parts.append(f"where  k.cal_date = 20250331 and (k.report_flag_tprt = '1' or k.report_flag_east5 = '1')  and   proj.busi_scope = '1' and  ")
    sql_parts.append(f" -- DM限定条件")
    sql_parts.append(f"  {null_check}")
    
    # 参数过滤条件
    for pf in param_filters:
        sql_parts.append(pf)
    
    rebuilt_sql = '\n'.join(sql_parts)
    
    # 添加注释头
    param_labels = []
    for code, value in param_combo.items():
        param_labels.append(build_param_value_label(code, value, param_config))
    
    header_lines = []
    header_lines.append(f"--报表字段：{check_info.report_table}.{check_info.report_field}")
    header_lines.append(f"--DM字段：{dm_table}.{dm_field}")
    header_lines.append(f"-- 参数组合: {', '.join(param_labels)}")
    header_lines.append(f"-- 检查编号: {check_info.check_code}")
    header_lines.append(f"-- 勾稽名称: {check_name}")
    header = '\n'.join(header_lines) + '\n'
    
    rebuilt_sql = header + rebuilt_sql
    
    if logger:
        logger.info(f"  血缘重建: {lineage_key} 参数组合 {param_combo} → STG={stg_table}.{stg_col_name}({stg_col_name_cn})")
    
    return rebuilt_sql


def _get_stg_table_cn(stg_table: str) -> str:
    """根据STG表名返回中文名"""
    _STG_TABLE_CN = {
        'T2_TCMP_PROJECTINFO': '项目信息',
        'T2_TCMP_EMPLOYEEINFO': '员工信息表',
        'T2_TCMP_DEPARTMENTINFO': '部门信息',
    }
    return _STG_TABLE_CN.get(stg_table, stg_table)


def _build_sql_header(check_info: ReportCheckInfo, insp_row: Optional[InspectionRow] = None,
                     param_combo: Optional[Dict[str, str]] = None,
                     param_config: Optional[Dict[str, ParamDefinition]] = None) -> str:
    """
    生成SQL文件的统一注释表头。

    包含：报表字段、DM字段、检查编号、勾稽名称。
    v0.7新增：参数组合标注。

    Args:
        check_info: 检查项信息
        insp_row: 巡检列匹配行（可选，占位SQL时为None）
        param_combo: 参数值组合（可选，展开SQL时使用）
        param_config: 参数配置字典（可选，展开SQL时使用）

    Returns:
        注释表头字符串
    """
    report_table = check_info.report_table.strip()
    report_field = check_info.report_field.strip()
    dm_table = insp_row.dm_table if insp_row else ''
    dm_field = insp_row.dm_field if insp_row else ''

    header_lines = []
    header_lines.append(f"--报表字段：{report_table}.{report_field}")
    header_lines.append(f"--DM字段：{dm_table}.{dm_field}")
    
    # v0.7: 参数组合标注
    if param_combo:
        param_labels = []
        for code, value in param_combo.items():
            label = build_param_value_label(code, value, param_config or {})
            param_labels.append(label)
        header_lines.append(f"-- 参数组合: {', '.join(param_labels)}")
    
    header_lines.append(f"-- 检查编号: {check_info.check_code}")
    header_lines.append(f"-- 勾稽名称: {check_info.check_name}")

    return '\n'.join(header_lines) + '\n'


def _remove_business_param_conditions(sql: str, business_param_codes: List[str]) -> str:
    """
    从SQL中移除业务参数的URP_PARAM_CONFIG条件。
    
    业务参数已通过参数展开为独立SQL，不需要在WHERE中保留。
    个性化参数的URP_PARAM_CONFIG查询保持不变。
    
    Args:
        sql: SQL字符串
        business_param_codes: 需要移除的业务参数代码列表
        
    Returns:
        处理后的SQL字符串
    """
    for param_code in business_param_codes:
        # 模式1: AND (select param_value from urp3.URP_PARAM_CONFIG t/pc where param_code='XXX' and status='1') = 'Y'
        pattern1 = re.compile(
            r"\s+AND\s+\(?\s*select\s+param_value\s+from\s+urp3\.URP_PARAM_CONFIG\s+\w+\s+"
            r"where\s+param_code\s*=\s*'" + re.escape(param_code) + r"'\s+"
            r"and\s+status\s*=\s*'1'\s*\)?\s*=\s*'[^']*'",
            re.IGNORECASE
        )
        sql = pattern1.sub('', sql)
        
        # 模式2: AND (select param_value from urp3.URP_PARAM_CONFIG t where param_code='XXX' and status='1') IS NOT NULL
        pattern2 = re.compile(
            r"\s+AND\s+\(\s*select\s+param_value\s+from\s+urp3\.URP_PARAM_CONFIG\s+\w+\s+"
            r"where\s+param_code\s*=\s*'" + re.escape(param_code) + r"'\s+"
            r"and\s+status\s*=\s*'1'\s*\)\s+IS\s+NOT\s+NULL",
            re.IGNORECASE
        )
        sql = pattern2.sub('', sql)
        
        # 模式3: or (select ... param_code='XXX' ... ) = 'Y' (带or的情况)
        pattern3 = re.compile(
            r"\s+or\s+\(?\s*select\s+param_value\s+from\s+urp3\.URP_PARAM_CONFIG\s+\w+\s+"
            r"where\s+param_code\s*=\s*'" + re.escape(param_code) + r"'\s+"
            r"and\s+status\s*=\s*'1'\s*\)?\s*=\s*'[^']*'",
            re.IGNORECASE
        )
        sql = pattern3.sub('', sql)
    
    # 清理多余空行（移除条件后可能留下空行）
    sql = re.sub(r'\n\s*\n\s*\n', '\n\n', sql)
    
    return sql


def generate_param_expanded_sqls(
    sql: str,
    check_info: ReportCheckInfo,
    insp_row: Optional[InspectionRow],
    param_config: Dict[str, ParamDefinition],
    stg_key: str,
    stg_table_info: dict,
    has_project: bool,
    logger=None,
    field_param_codes: Optional[List[str]] = None
) -> List[Dict]:
    """
    对包含业务参数的SQL进行参数展开，生成多份SQL。
    
    对于每条检查项：
    1. 从SQL模板中提取参数引用
    2. 分类为 business 和 personalization
    3. 对 business 参数计算笛卡尔积
    4. 每个参数值组合生成一份SQL
    5. 移除 business 参数的 URP_PARAM_CONFIG 条件
    6. 保留 personalization 参数的 URP_PARAM_CONFIG 条件
    7. 添加参数组合标注
    
    如果SQL中没有业务参数，返回单元素的列表（原SQL不变）。
    
    v0.8: 支持 field_param_codes 参数 — 当SQL模板中无URP_PARAM_CONFIG引用，
    但通过 param_field_mapping.xlsx 发现字段受业务参数控制时，使用外部传入的参数代码进行展开。
    
    Args:
        sql: 原始SQL字符串
        check_info: 检查项信息
        insp_row: 巡检列匹配行
        param_config: 参数配置字典
        stg_key: STG主键表达式
        stg_table_info: STG表字段中文名映射
        has_project: 是否与项目相关
        logger: 日志记录器
        field_param_codes: 从字段级映射获取的参数代码列表（v0.8新增）
        
    Returns:
        list: [{'sql': str, 'param_combo': dict, 'param_codes': list, 'param_labels': list}, ...]
    """
    # 1. 提取SQL中的参数引用
    all_codes = get_all_param_codes(sql)
    
    # v0.8: 合并字段级参数映射的代码
    if field_param_codes:
        for code in field_param_codes:
            if code not in all_codes:
                all_codes.append(code)
    
    if not all_codes:
        # 无参数引用，返回原SQL
        return [{'sql': sql, 'param_combo': {}, 'param_codes': [], 'param_labels': []}]
    
    # 2. 分类参数
    business_codes, personal_codes = classify_params_in_sql(sql, param_config)
    
    # v0.8: 将字段级映射的参数代码加入 business_codes
    if field_param_codes:
        for code in field_param_codes:
            if code not in business_codes and code not in personal_codes:
                business_codes.append(code)
    
    # 2b. 将不在配置中的业务参数重新分类为个性化参数（因为没有值列表无法展开）
    expandable_business = []
    for code in business_codes:
        actual_code = code
        if code not in param_config:
            equiv = _get_equivalent_param(code)
            if equiv and equiv in param_config:
                actual_code = equiv
        if actual_code in param_config and param_config[actual_code].param_values:
            expandable_business.append(code)
        else:
            # 不在配置中或没有值列表 → 无法展开，视为个性化参数
            personal_codes.append(code)
    business_codes = expandable_business
    
    if not business_codes:
        # 只有个性化参数（或无法展开的业务参数），不需要展开
        return [{'sql': sql, 'param_combo': {}, 'param_codes': personal_codes, 'param_labels': []}]
    
    # 3. 计算业务参数的笛卡尔积
    combos = expand_business_params(param_config, business_codes)
    
    if len(combos) <= 1:
        # 只有一种组合（可能是默认值），不需要展开
        combo = combos[0] if combos else {}
        # 但仍需添加参数标注
        if combo:
            param_labels = []
            for code, value in combo.items():
                param_labels.append(build_param_value_label(code, value, param_config))
            return [{'sql': sql, 'param_combo': combo, 'param_codes': list(set(business_codes + personal_codes)), 
                     'param_labels': param_labels}]
        return [{'sql': sql, 'param_combo': {}, 'param_codes': list(set(business_codes + personal_codes)), 'param_labels': []}]
    
    # 4. 对每个参数值组合生成SQL
    results = []
    for combo in combos:
        # 复制SQL并替换业务参数条件
        expanded_sql = sql
        
        # v0.7.1: 先替换CASE WHEN参数（将CASE WHEN块替换为对应值的THEN表达式）
        expanded_sql = _replace_case_when_params(expanded_sql, combo, param_config)
        
        # 移除业务参数的URP_PARAM_CONFIG条件（简单AND/OR条件）
        expanded_sql = _remove_business_param_conditions(expanded_sql, business_codes)
        
        # v0.7.1: 进一步清理残留的业务参数行内条件
        expanded_sql = _remove_inline_param_conditions(expanded_sql, business_codes)
        
        # 构建参数标注
        param_labels = []
        for code, value in combo.items():
            param_labels.append(build_param_value_label(code, value, param_config))
        
        # 替换SQL头部（如果有参数标注）
        original_header = _build_sql_header(check_info, insp_row)
        new_header = _build_sql_header(check_info, insp_row, combo, param_config)
        
        if param_labels:
            param_comment = f"-- 参数组合: {', '.join(param_labels)}"
            if '-- 参数组合:' in expanded_sql:
                expanded_sql = re.sub(r'-- 参数组合:.*', param_comment, expanded_sql, count=1)
            elif '-- 检查编号:' in expanded_sql:
                expanded_sql = expanded_sql.replace('-- 检查编号:', param_comment + '\n-- 检查编号:', 1)
            elif expanded_sql.startswith('--'):
                lines = expanded_sql.split('\n', 1)
                expanded_sql = lines[0] + '\n' + param_comment + ('\n' + lines[1] if len(lines) > 1 else '')
            else:
                expanded_sql = param_comment + '\n' + expanded_sql
        
        results.append({
            'sql': expanded_sql,
            'param_combo': combo,
            'param_codes': list(set(business_codes + personal_codes)),
            'param_labels': param_labels
        })
    
    if logger:
        logger.info(f"检查项 {check_info.check_code} 参数展开: {len(business_codes)} 个业务参数, "
                    f"{len(personal_codes)} 个个性化参数, 生成 {len(results)} 份SQL")
    
    return results


def _replace_case_when_params(sql: str, param_combo: Dict[str, str],
                              param_config: Dict[str, ParamDefinition]) -> str:
    """
    对SQL中的CASE WHEN参数表达式进行值替换。

    当业务参数出现在CASE WHEN中时，将整个 CASE ... END 块替换为
    当前参数值对应的THEN分支表达式。

    例如：当 SRDT5_YGH=1 时：
      CASE
        WHEN (select ... param_code='SRDT5_YGH' ...) = 1 THEN k.LOGIN_ALIAS
        WHEN (select ... param_code='SRDT5_YGH' ...) = 2 THEN k.EMPLOYEE_ID
        ...
      END
    替换为：
      k.LOGIN_ALIAS

    Args:
        sql: SQL字符串
        param_combo: 当前参数值组合 {param_code: value}
        param_config: 参数配置字典

    Returns:
        替换后的SQL字符串
    """
    from parsers.param_config_loader import _get_equivalent_param

    for param_code, param_value in param_combo.items():
        # 查找等价参数代码
        equiv_code = _get_equivalent_param(param_code) or param_code

        # 匹配 CASE ... END 块中包含 param_code 的模式
        # 模式: CASE WHEN (select ... param_code='XXX' ...) = N THEN expr [WHEN ... THEN ...] END
        case_pattern = re.compile(
            r'CASE\s+((?:WHEN\s+\(?\s*select\s+param_value\s+from\s+urp3\.URP_PARAM_CONFIG\s+\w+\s+'
            r'where\s+param_code\s*=\s*\'' + re.escape(param_code) + r'\'' +
            r'\s+and\s+status\s*=\s*\'1\'\s*\)?\s*=\s*\d+\s+THEN\s+[^\n]+'
            r'(?:\s+WHEN\s+\(?\s*select\s+param_value\s+from\s+urp3\.URP_PARAM_CONFIG\s+\w+\s+'
            r'where\s+param_code\s*=\s*\'' + re.escape(param_code) + r'\'' +
            r'\s+and\s+status\s*=\s*\'1\'\s*\)?\s*=\s*\d+\s+THEN\s+[^\n]+)*)'
            r')\s+END',
            re.IGNORECASE | re.DOTALL
        )

        def _replace_case_block(match):
            case_body = match.group(1)
            # 从CASE WHEN分支中找到匹配当前参数值的THEN表达式
            # 匹配模式: WHEN (...) = <value> THEN <expression>
            branch_pattern = re.compile(
                r'WHEN\s+\(?\s*select\s+param_value\s+from\s+urp3\.URP_PARAM_CONFIG\s+\w+\s+'
                r'where\s+param_code\s*=\s*\'' + re.escape(param_code) + r'\'' +
                r'\s+and\s+status\s*=\s*\'1\'\s*\)?\s*=\s*(\d+)\s+THEN\s+([^\n]+)',
                re.IGNORECASE
            )

            for branch_match in branch_pattern.finditer(case_body):
                branch_value = branch_match.group(1).strip()
                branch_expr = branch_match.group(2).strip()

                # 比较分支值与当前参数值
                try:
                    if str(branch_value) == str(param_value):
                        return branch_expr
                except (ValueError, TypeError):
                    pass

            # 如果没有匹配到，使用ELSE分支（如果有）或保留原CASE
            # 尝试查找ELSE分支
            else_pattern = re.compile(
                r'ELSE\s+([^\n]+?)(?:\s+END)',
                re.IGNORECASE
            )
            else_match = else_pattern.search(case_body)
            if else_match:
                return else_match.group(1).strip()

            # 无法匹配，保留整个CASE WHEN块不变
            return match.group(0)

        sql = case_pattern.sub(_replace_case_block, sql)

    return sql


def _remove_inline_param_conditions(sql: str, business_param_codes: List[str]) -> str:
    """
    移除SQL中业务参数的行内URP_PARAM_CONFIG条件（非CASE WHEN模式）。

    处理模式包括：
    1. AND (select param_value from urp3.URP_PARAM_CONFIG ... where param_code='XXX' ...) = 'Y'
    2. AND (select ... param_code='XXX' ...) IS NOT NULL
    3. OR (select ... param_code='XXX' ...) = 'Y'
    4. AND (select ... param_code='XXX' ...) <> 'VALUE'   (不等于某值)
    5. 整行包含 param_code='XXX' 且属于业务参数的条件

    对于包裹在 AND(...) 或嵌套条件中的业务参数行，逐行移除整行。

    Args:
        sql: SQL字符串
        business_param_codes: 需要移除的业务参数代码列表

    Returns:
        处理后的SQL字符串
    """
    for param_code in business_param_codes:
        # 模式1: AND (select ... param_code='XXX' ...) = 'Y'
        pattern1 = re.compile(
            r"\s+AND\s+\(?\s*select\s+param_value\s+from\s+urp3\.URP_PARAM_CONFIG\s+\w+\s+"
            r"where\s+param_code\s*=\s*'" + re.escape(param_code) + r"'\s+"
            r"and\s+status\s*=\s*'1'\s*\)?\s*=\s*'[^']*'",
            re.IGNORECASE
        )
        sql = pattern1.sub('', sql)

        # 模式2: AND (select ... param_code='XXX' ...) IS NOT NULL
        pattern2 = re.compile(
            r"\s+AND\s+\(\s*select\s+param_value\s+from\s+urp3\.URP_PARAM_CONFIG\s+\w+\s+"
            r"where\s+param_code\s*=\s*'" + re.escape(param_code) + r"'\s+"
            r"and\s+status\s*=\s*'1'\s*\)\s+IS\s+NOT\s+NULL",
            re.IGNORECASE
        )
        sql = pattern2.sub('', sql)

        # 模式3: OR (select ... param_code='XXX' ...) = 'Y'
        pattern3 = re.compile(
            r"\s+or\s+\(?\s*select\s+param_value\s+from\s+urp3\.URP_PARAM_CONFIG\s+\w+\s+"
            r"where\s+param_code\s*=\s*'" + re.escape(param_code) + r"'\s+"
            r"and\s+status\s*=\s*'1'\s*\)?\s*=\s*'[^']*'",
            re.IGNORECASE
        )
        sql = pattern3.sub('', sql)

    # 逐行清理：移除整个行含业务参数且不是CASE WHEN的行
    lines = sql.split('\n')
    cleaned_lines = []
    for line in lines:
        # 检查该行是否含有业务参数的URP_PARAM_CONFIG引用
        line_has_business_param = False
        is_case_when_line = re.search(r'\bCASE\b', line, re.IGNORECASE) or \
                            re.search(r'\bWHEN\b.*\bTHEN\b', line, re.IGNORECASE)

        for param_code in business_param_codes:
            if f"param_code='{param_code}'" in line or \
               f"param_code = '{param_code}'" in line.lower():
                line_has_business_param = True
                break

        if line_has_business_param and not is_case_when_line:
            # 额外检查：该行是否是多行CASE WHEN的一部分
            # 如果该行以AND/or开头且包含业务参数的条件，直接跳过
            stripped = line.strip()
            if stripped.upper().startswith('AND') or stripped.upper().startswith('OR'):
                # 检查是否只是简单条件（不包含CASE WHEN）
                if 'CASE' not in line.upper():
                    continue  # 跳过此行
            elif re.match(r'\s*\(?\s*select\s+param_value\s+from\s+urp3', stripped, re.IGNORECASE):
                # 独立的子查询条件行
                continue

        cleaned_lines.append(line)

    sql = '\n'.join(cleaned_lines)

    # 清理多余空行（移除条件后可能留下空行）
    sql = re.sub(r'\n\s*\n\s*\n', '\n\n', sql)

    return sql


def _replace_project_fields(sql: str) -> str:
    """将项目辅助字段替换为 NULL"""
    sql = sql.replace(
        "proj.count_proj_code,\n       proj.count_proj_name,",
        "NULL count_proj_code,\n       NULL count_proj_name,"
    )
    # 替换 trust_manager 等
    for field in ['trust_manager', 'deal_manage', 'trust_manager_b', 'trust_manager_c', 'beit_dept']:
        sql = re.sub(
            rf"urp3_xtdw\.fn_company_dict\(proj\.{field},\s*'[12]'\)\s+{field}",
            f"NULL {field}",
            sql
        )
    # 移除 proj 的 INNER JOIN
    # 保持 DM 表 JOIN 不变，但 proj 表有时在模板中 INNER JOIN
    # 这里暂不复杂化，非项目报表的模板通常已不同于项目报表

    return sql


def _assemble_sql(check_info: ReportCheckInfo,
                  insp_row: InspectionRow,
                  stg_key: str,
                  stg_table_info: dict,
                  has_project: bool,
                  logger=None) -> str:
    """
    无 SQL 模板时，基于 InspectionRow 信息拼装 SQL。
    """
    dm_table = insp_row.dm_table
    dm_field = insp_row.dm_field or check_info.report_field
    stg_table = insp_row.stg_table
    stg_field = insp_row.stg_field or check_info.report_field
    stg_cn = stg_table_info.get(stg_table, {})
    stg_table_cn = insp_row.stg_table_cn or stg_table
    stg_field_cn = insp_row.stg_field_cn or stg_field

    # 构建 SELECT（每个单项不含尾部逗号，由 join 添加）
    select_parts = []
    select_parts.append(f"       '{check_info.check_name}' check_name")
    select_parts.append(f"       '{stg_table}' stg_table_name")
    select_parts.append(f"       '{stg_table_cn}' stg_table_name_cn")
    select_parts.append(f"       '{stg_field}' stg_col_name")
    select_parts.append(f"       '{stg_field_cn}' stg_col_name_cn")
    select_parts.append(f"       t.{stg_field} stg_col_value")

    if stg_key:
        select_parts.append(f"       {stg_key} stg_key")
    else:
        select_parts.append(f"       '项目编号['||t.project_code||']' stg_key")

    # 拼接SELECT子句：单项字段之间用逗号连接，多行块（已含逗号）直接追加
    sql = _build_sql_header(check_info, insp_row)
    sql += "SELECT DISTINCT\n"
    sql += ",\n".join(select_parts) + ",\n"
    if has_project:
        sql += BUSINESS_FIELDS + "\n"
    else:
        sql += NO_PROJECT_FIELDS + "\n"
    sql += REPORT_FLAG_FIELDS + "\n"

    # 构建 FROM
    from_parts = []
    from_parts.append(f"FROM {dm_table} k")

    # 如果有项目，JOIN proj 表
    if has_project:
        from_parts.append(f"INNER JOIN urp3_xtdw.dw_d_count_proj_info proj")
        from_parts.append(f"    ON k.count_proj_code = proj.count_proj_code")
        if insp_row.stg_table.startswith('T2_') or insp_row.stg_table.startswith('TS_'):
            stg_schema = "urp3_xtstg"
            from_parts.append(f"INNER JOIN {stg_schema}.{stg_table} t")
            from_parts.append(f"    ON proj.proj_code = t.project_code")

    # 构建 WHERE
    where_parts = []
    where_parts.append(f"WHERE k.cal_date = ${{BUSI_DATE}}")
    where_parts.append(f"  AND (k.report_flag_tprt = '1' OR k.report_flag_east5 = '1')")

    if has_project:
        where_parts.append(f"  AND proj.busi_scope = '1'")

    # 必填/条件必填检查
    if check_info.is_conditional and insp_row.extra_condition:
        where_parts.append(f"  AND {insp_row.extra_condition}")
    elif check_info.is_conditional:
        extra = build_extra_condition(check_info, "k")
        if extra:
            where_parts.append(extra)

    where_parts.append(f"  AND trim(k.{dm_field}) IS NULL")

    # 组装 SQL（SELECT部分已在上方的select_parts中构建）
    sql += "\n".join(from_parts) + "\n"
    sql += "\n".join(where_parts) + ";"

    return sql


def generate_placeholder_sql(check_info: ReportCheckInfo, reason: str, logger=None) -> str:
    """为未找到映射的检查项生成占位 SQL"""
    header = _build_sql_header(check_info)
    sql = f"""{header}
-- !!! 未自动生成：{reason} !!!
-- TODO: 请人工补充以下巡检SQL

SELECT DISTINCT
       '{check_info.check_name}' check_name,
       'TODO' stg_table_name,
       'TODO' stg_table_name_cn,
       'TODO' stg_col_name,
       'TODO' stg_col_name_cn,
       'TODO' stg_col_value,
       'TODO' stg_key,
       NULL count_proj_code,
       NULL count_proj_name,
       NULL trust_manager,
       NULL deal_manage,
       NULL trust_manager_b,
       NULL trust_manager_c,
       NULL beit_dept,
       k.report_flag_tprt,
       k.report_flag_east5
FROM TODO_DM_TABLE k
WHERE k.cal_date = ${{BUSI_DATE}}
  AND (k.report_flag_tprt = '1' OR k.report_flag_east5 = '1')
  AND trim(k.TODO_FIELD) IS NULL
;
"""
    if logger:
        logger.warn(f"检查项 {check_info.check_code} 未生成SQL: {reason}", Logger.WARN_MAPPING_MISSING)

    return sql


def generate_sql_for_check(check_info: ReportCheckInfo,
                           mapping: ETLMapping,
                           stg_key_map: dict,
                           stg_table_info: dict,
                           logger=None) -> str:
    """
    为单条检查项生成完整的巡检 SQL。

    逻辑：
    1. 通过 find_mapping_for_check 查找巡检列映射
    2. 找到 → 使用巡检列的 SQL 模板或拼装
    3. 未找到 → 生成占位 SQL

    Args:
        check_info: 检查项信息
        mapping: ETLMapping 实例
        stg_key_map: STG 业务主键映射（stg_key_list.xlsx）
        stg_table_info: STG 表字段中文名映射
        logger: 日志记录器

    Returns:
        生成的 SQL 字符串
    """
    # 查找巡检列映射
    insp_row = find_mapping_for_check(check_info, mapping, logger)

    if insp_row is None:
        return generate_placeholder_sql(check_info, "未找到巡检列映射", logger)

    # 获取 STG 主键
    stg_key = mapping.stg_key_map.get(insp_row.stg_table, "")
    if not stg_key:
        stg_key = stg_key_map.get(insp_row.stg_table, "")

    # 使用巡检列映射生成 SQL
    sql = generate_sql_from_template(check_info, insp_row, stg_key, stg_table_info, logger)

    return sql


def load_existing_mapping(output_dir: str) -> dict:
    """
    读取现有的映射表CSV，返回 {check_code: suc_number} 映射。
    
    如果映射表不存在，返回空字典（将按顺序生成新编号）。
    如果映射表存在，已有的check_code保持原SUC编号不变。
    """
    import csv
    csv_path = os.path.join(output_dir, "check_suc_mapping.csv")
    existing = {}
    if os.path.exists(csv_path):
        try:
            with open(csv_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    existing[row['check_code']] = row['suc_number']
        except Exception:
            pass  # 读取失败则返回空字典
    return existing


def write_mapping_md(mapping_records: list, output_dir: str, logger=None):
    """
    将映射关系写入 markdown 文件。
    
    文件格式：markdown 表格，包含 SUC编号、勾稽代码、勾稽名称、匹配类型等。
    """
    md_path = os.path.join(output_dir, "check_suc_mapping.md")
    
    # 统计
    total = len(mapping_records)
    exact_template = sum(1 for r in mapping_records if r['match_type'] == 'EXACT_TEMPLATE')
    exact_assemble = sum(1 for r in mapping_records if r['match_type'] == 'EXACT_ASSEMBLE')
    dm_supplement_template = sum(1 for r in mapping_records if r['match_type'] == 'DM_SUPPLEMENT_TEMPLATE')
    dm_supplement_assemble = sum(1 for r in mapping_records if r['match_type'] == 'DM_SUPPLEMENT_ASSEMBLE')
    placeholder = sum(1 for r in mapping_records if r['match_type'] == 'PLACEHOLDER')
    
    lines = []
    lines.append("# 勾稽代码 ↔ SUC编号 映射表")
    lines.append("")
    lines.append(f"> 自动生成于 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"> 总计 {total} 条：精确匹配(模板) {exact_template} 条 | 精确匹配(拼装) {exact_assemble} 条 | DM补充(模板) {dm_supplement_template} 条 | DM补充(拼装) {dm_supplement_assemble} 条 | 需人工 {placeholder} 条")
    lines.append(f"> 覆盖率: {(total - placeholder) / total * 100:.1f}%")
    lines.append("")
    lines.append("| SUC编号 | 勾稽代码 | 勾稽名称 | 报表名 | 报表字段 | 必填类型 | 匹配类型 | DM表 | DM字段 | STG表 | STG字段 | SQL文件 | 参数代码 | 参数值 |")
    lines.append("|---------|----------|----------|--------|----------|----------|----------|------|--------|--------|--------|---------|----------|--------|")
    
    for r in mapping_records:
        # 截断过长的check_name
        short_name = r['check_name'][:40] + '...' if len(r['check_name']) > 40 else r['check_name']
        is_mandatory_str = '必填' if '必填' in r['is_mandatory'] or r['is_mandatory'] == '必填' else '条件必填'
        
        match_type_label = {
            'EXACT_TEMPLATE': '✅ 模板',
            'EXACT_ASSEMBLE': '✅ 拼装',
            'DM_SUPPLEMENT_TEMPLATE': '⚠️ DM补充(模板)',
            'DM_SUPPLEMENT_ASSEMBLE': '⚠️ DM补充(拼装)',
            'PLACEHOLDER': '❌ 需人工',
        }.get(r['match_type'], r['match_type'])
        
        param_codes_str = r.get('param_codes', '')
        param_values_str = r.get('param_values', '')
        lines.append(
            f"| {r['suc_number']} | {r['check_code']} | {short_name} | "
            f"{r['report_table']} | {r['report_field']} | {is_mandatory_str} | "
            f"{match_type_label} | {r['dm_table']} | {r['dm_field']} | "
            f"{r['stg_table']} | {r['stg_field']} | {r['sql_file']} | "
            f"{param_codes_str} | {param_values_str} |"
        )
    
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 匹配类型说明")
    lines.append("")
    lines.append("| 匹配类型 | 说明 |")
    lines.append("|----------|------|")
    lines.append("| ✅ 模板 | 巡检列精确匹配，且SQL模板可用。字段和关联条件均正确。 |")
    lines.append("| ✅ 拼装 | 巡检列精确匹配，但无SQL模板，从映射信息拼装。字段正确但JOIN条件为通用模式。 |")
    lines.append("| ⚠️ DM补充(模板) | 通过DM字段映射表补充，借用同DM表的模板（保留JOIN条件和data_source过滤），需人工确认STG字段。 |")
    lines.append("| ⚠️ DM补充(拼装) | 通过DM字段映射表补充且无可用模板，使用通用拼装。JOIN条件和字段均需人工确认。 |")
    lines.append("| ❌ 需人工 | 未找到任何映射，需人工补充完整的巡检SQL。 |")
    lines.append("")
    lines.append("## 更新规则")
    lines.append("")
    lines.append("1. 已有的勾稽代码保持原SUC编号不变")
    lines.append("2. 新增的勾稽代码追加到末尾，分配新SUC编号")
    lines.append("3. 删除的勾稽代码保留其SUC编号但标记为已删除")
    
    with open(md_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    
    if logger:
        logger.info(f"映射表MD已写入: {md_path}")
    
    print(f"映射表(MD): {md_path}")


def write_mapping_csv(mapping_records: list, output_dir: str, logger=None):
    """
    将映射关系写入 CSV 文件，便于程序化读取和Excel打开。
    """
    import csv
    csv_path = os.path.join(output_dir, "check_suc_mapping.csv")
    
    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=[
            'suc_number', 'check_code', 'check_name', 'report_table', 'report_field',
            'is_mandatory', 'match_type', 'dm_table', 'dm_field',
            'stg_table', 'stg_field', 'sql_file', 'param_codes', 'param_values'
        ])
        writer.writeheader()
        writer.writerows(mapping_records)
    
    if logger:
        logger.info(f"映射表CSV已写入: {csv_path}")
    
    print(f"映射表(CSV): {csv_path}")


def main():
    """主执行入口（v2 + v0.7参数展开）"""
    # 路径配置
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    report_check_path = os.path.join(base_dir, "report_check_list.xlsx")
    my_poor_solution_path = os.path.join(base_dir, "my_poor_solution.xlsx")
    stg_key_path = os.path.join(base_dir, "stg_key_list.xlsx")
    table_structure_path = os.path.join(base_dir, "table_structure_list.xlsx")
    param_config_path = os.path.join(base_dir, "code_param_list.xlsx")
    field_param_mapping_path = os.path.join(base_dir, "param_field_mapping.xlsx")
    lineage_config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "etl_lineage_config.json")
    output_dir = os.path.join(base_dir, "output_stg_file")
    log_path = os.path.join(output_dir, "generation_log.txt")

    # 初始化日志
    os.makedirs(output_dir, exist_ok=True)
    logger = Logger(log_path)
    logger.section("第1步：读取输入文件")

    # 1. 读取报表检查清单
    logger.info(f"读取报表检查清单: {report_check_path}")
    checks = load_report_checks(report_check_path)
    logger.info(f"共 {len(checks)} 条检查项")

    # 2. 加载映射（主映射源 = 巡检列 sheet）
    logger.info(f"加载映射: {my_poor_solution_path}")
    mapping = load_my_poor_solution(my_poor_solution_path, logger)

    # 2d. 加载参数配置（v0.7新增）
    param_config = {}
    if os.path.exists(param_config_path):
        logger.info(f"加载参数配置: {param_config_path}")
        param_config = load_param_config(param_config_path, logger)
        business_count = sum(1 for p in param_config.values() if p.category == 'business')
        personal_count = sum(1 for p in param_config.values() if p.category == 'personalization')
        logger.info(f"  业务参数: {business_count} 个, 个性化参数: {personal_count} 个")
    else:
        logger.warn(f"参数配置文件不存在: {param_config_path}, 跳过参数展开")
    
    # 2e. 加载字段级参数映射（v0.8新增）
    field_param_map = {}
    if param_config and os.path.exists(field_param_mapping_path):
        logger.info(f"加载字段级参数映射: {field_param_mapping_path}")
        field_param_map = load_field_param_mapping(field_param_mapping_path, param_config, logger)
    else:
        if logger:
            logger.info(f"字段级参数映射文件不存在或无参数配置，跳过")
    
    # 2f. 加载ETL血缘配置（v0.9新增）
    lineage_config = {"dm_field_lineage": {}, "dw_stg_mapping": {}}
    if param_config and os.path.exists(lineage_config_path):
        logger.info(f"加载ETL血缘配置: {lineage_config_path}")
        lineage_config = load_etl_lineage_config(lineage_config_path, logger)
    else:
        if logger:
            logger.info(f"ETL血缘配置文件不存在，跳过血缘重建")

    # 3. 加载 STG 业务主键
    logger.info(f"加载 STG 业务主键: {stg_key_path}")
    stg_key_map = load_stg_key_list(stg_key_path)
    logger.info(f"STG 业务主键数: {len(stg_key_map)}")

    # 4. 加载 STG 表字段信息
    logger.info(f"加载 STG 表字段信息: {table_structure_path}")
    stg_table_info = load_stg_table_info(table_structure_path)
    logger.info(f"STG 表信息数: {len(stg_table_info)}")

    # 5. 生成 SQL（含参数展开）
    logger.section("第2步：生成巡检SQL（含参数展开）")
    # sql_list 改为支持多份SQL的展开列表
    # 每个元素: (check_code, sql, param_combo, param_codes, param_labels)
    sql_list = []
    param_expanded_count = 0
    param_no_expand_count = 0
    template_count = 0
    assembled_count = 0
    placeholder_count = 0

    for i, check in enumerate(checks, start=1):
        insp_row = find_mapping_for_check(check, mapping, logger)
        stg_key = ""
        if insp_row:
            stg_key = mapping.stg_key_map.get(insp_row.stg_table, "")
            if not stg_key:
                stg_key = stg_key_map.get(insp_row.stg_table, "")
        
        # 生成基础SQL
        if insp_row is None:
            base_sql = generate_placeholder_sql(check, "未找到巡检列映射", logger)
            match_type = "PLACEHOLDER"
        else:
            base_sql = generate_sql_from_template(check, insp_row, stg_key, stg_table_info, logger)
            # 判断匹配类型
            field_alias = f"{check.report_table}.{check.report_field}"
            if field_alias in mapping.inspection_map:
                rows = mapping.inspection_map[field_alias]
                insp = rows[0]
                if insp.is_from_dm_supplement:
                    match_type = "DM_SUPPLEMENT_TEMPLATE" if insp.sql_template else "DM_SUPPLEMENT_ASSEMBLE"
                elif insp.sql_template:
                    match_type = "EXACT_TEMPLATE"
                else:
                    match_type = "EXACT_ASSEMBLE"
            else:
                if insp_row and insp_row.is_from_dm_supplement:
                    match_type = "DM_SUPPLEMENT_TEMPLATE" if insp_row.sql_template else "DM_SUPPLEMENT_ASSEMBLE"
                else:
                    match_type = "DM_SUPPLEMENT_ASSEMBLE"
        
        # v0.9: 参数展开 — 优先使用血缘重建
        # 1) 检查血缘配置中是否有该DM字段的完整映射
        # 2) 如有，使用血缘重建生成结构不同的SQL
        # 3) 如无，退回v0.8逻辑（WHERE条件替换）
        has_sql_param_ref = 'URP_PARAM_CONFIG' in base_sql
        has_template_param = insp_row and insp_row.sql_template and '#[' in insp_row.sql_template
        
        # v0.9: 检查血缘配置
        lineage_key = ''
        has_lineage = False
        if param_config and insp_row and lineage_config.get('dm_field_lineage'):
            dm_table_l = (insp_row.dm_table or '').strip().upper()
            dm_field_l = (check.report_field or '').strip().upper()
            lineage_key = f"{dm_table_l}.{dm_field_l}"
            has_lineage = lineage_key in lineage_config['dm_field_lineage']
        
        field_param_codes = None
        if param_config and insp_row and not has_sql_param_ref and not has_template_param and not has_lineage:
            dm_table = (insp_row.dm_table or '').strip().upper()
            dm_field = (check.report_field or '').strip().upper()
            lookup_key = (dm_table, dm_field)
            if lookup_key in field_param_map:
                field_param_codes = field_param_map[lookup_key]
                if logger:
                    logger.info(f"  检查项 {check.check_code} 字段 {dm_table}.{dm_field} "
                                f"受业务参数控制: {field_param_codes}（来自字段级映射）")
        
        if param_config and insp_row and has_lineage:
            # v0.9: 血缘重建路径 — 参数驱动SQL结构重建
            field_lineage = lineage_config['dm_field_lineage'][lineage_key]
            lineage_params = field_lineage.get('params', {})
            
            # 确定参与展开的参数代码
            lineage_business_codes = list(lineage_params.keys())
            
            # 计算笛卡尔积
            combos = expand_business_params(param_config, lineage_business_codes)
            
            if len(combos) > 1:
                lineage_rebuilt_count = 0
                for combo in combos:
                    rebuilt_sql = rebuild_sql_with_lineage(
                        base_sql=base_sql,
                        check_info=check,
                        insp_row=insp_row,
                        param_combo=combo,
                        param_config=param_config,
                        lineage_config=lineage_config,
                        stg_key=stg_key,
                        stg_table_info=stg_table_info,
                        has_project=is_project_related(check),
                        logger=logger
                    )
                    
                    if rebuilt_sql:
                        param_labels = []
                        for code, value in combo.items():
                            param_labels.append(build_param_value_label(code, value, param_config))
                        
                        sql_list.append((
                            check.check_code,
                            rebuilt_sql,
                            combo,
                            lineage_business_codes,
                            param_labels,
                            'LINEAGE_REBUILD',
                            insp_row
                        ))
                        lineage_rebuilt_count += 1
                    else:
                        if logger:
                            logger.warn(f"  血缘重建返回None: {lineage_key} combo={combo}")
                
                if lineage_rebuilt_count > 0:
                    param_expanded_count += lineage_rebuilt_count - 1
                    if logger:
                        logger.info(f"  检查项 {check.check_code} 血缘重建: 生成 {lineage_rebuilt_count} 份SQL")
            else:
                # 只有1种组合，仍用原SQL
                sql_list.append((
                    check.check_code, base_sql, {}, [], [],
                    match_type, insp_row
                ))
                param_no_expand_count += 1
        
        elif param_config and insp_row and (has_sql_param_ref or has_template_param or field_param_codes):
            # v0.8: 原有路径 — SQL模板含参数引用时，仅替换WHERE条件
            expanded = generate_param_expanded_sqls(
                sql=base_sql,
                check_info=check,
                insp_row=insp_row,
                param_config=param_config,
                stg_key=stg_key,
                stg_table_info=stg_table_info,
                has_project=is_project_related(check),
                logger=logger,
                field_param_codes=field_param_codes
            )
            
            if len(expanded) > 1:
                # 多份SQL（参数展开）
                for exp in expanded:
                    sql_list.append((
                        check.check_code,
                        exp['sql'],
                        exp['param_combo'],
                        exp['param_codes'],
                        exp['param_labels'],
                        match_type,
                        insp_row
                    ))
                param_expanded_count += len(expanded) - 1  # 减1因为原始也算
            else:
                # 单份SQL（仅个性化参数或无业务参数）
                exp = expanded[0]
                sql_list.append((
                    check.check_code,
                    exp['sql'],
                    exp.get('param_combo', {}),
                    exp.get('param_codes', []),
                    exp.get('param_labels', []),
                    match_type,
                    insp_row
                ))
                param_no_expand_count += 1
        else:
            # 无参数引用，原样添加
            sql_list.append((
                check.check_code,
                base_sql,
                {},  # 无参数组合
                [],  # 无参数代码
                [],  # 无参数标签
                match_type,
                insp_row
            ))
        
        # 统计
        if "-- !!! 未自动生成" in base_sql:
            placeholder_count += 1
        elif match_type == "EXACT_TEMPLATE":
            template_count += 1
        elif match_type == "EXACT_ASSEMBLE":
            assembled_count += 1

        if i % 100 == 0:
            logger.info(f"已处理 {i}/{len(checks)} 条检查项")

    generated_total = template_count + assembled_count
    logger.info(f"SQL 生成完成：共 {len(sql_list)} 份SQL（含参数展开）")
    logger.info(f"  使用SQL模板: {template_count} 条")
    logger.info(f"  拼装生成: {assembled_count} 条")
    logger.info(f"  占位(需人工补充): {placeholder_count} 条")
    logger.info(f"  参数展开新增: {param_expanded_count} 份SQL")
    logger.info(f"  参数未展开(仅个性化): {param_no_expand_count} 条")

    # 6. 写入文件
    logger.section("第3步：写入文件")
    writer = BatchWriter(output_dir, batch_size=100, logger=logger)

    # 6a. 计算总文件数并写入
    # sql_list 现在每个元素是 (check_code, sql, param_combo, param_codes, param_labels, match_type, insp_row)
    # 只取前两个字段传给 writer
    sql_write_list = [(item[0], item[1]) for item in sql_list]
    result = writer.write_batch(sql_write_list)

    # 6b. 生成映射表（含参数信息）
    logger.section("第4步：生成映射表")
    mapping_records = []
    for i, item in enumerate(sql_list, start=1):
        check_code, sql, param_combo, param_codes, param_labels, match_type, insp_row = item
        check_idx = next((j for j, c in enumerate(checks) if c.check_code == check_code), 0)
        check = checks[check_idx] if check_idx < len(checks) else checks[0]
        suc_number = f"SUC{i:04d}"
        
        # 参数组合标签
        param_label_str = ', '.join(param_labels) if param_labels else ''
        param_codes_str = ', '.join(param_codes) if param_codes else ''
        
        mapping_records.append({
            'suc_number': suc_number,
            'check_code': check_code,
            'check_name': check.check_name,
            'report_table': check.report_table,
            'report_field': check.report_field,
            'is_mandatory': check.is_mandatory,
            'match_type': match_type,
            'dm_table': insp_row.dm_table if insp_row else '',
            'dm_field': insp_row.dm_field if insp_row else '',
            'stg_table': insp_row.stg_table if insp_row else '',
            'stg_field': insp_row.stg_field if insp_row else '',
            'sql_file': f"{writer._get_folder_name(i)}/{suc_number}.sql",
            'param_codes': param_codes_str,
            'param_values': param_label_str,
        })

    # 写入 md 和 csv（需要更新write_mapping_csv/md以支持新字段）
    write_mapping_md(mapping_records, output_dir, logger)
    write_mapping_csv(mapping_records, output_dir, logger)

    # 7. 生成摘要
    mandatory_count = sum(1 for c in checks if not c.is_conditional)
    conditional_count = sum(1 for c in checks if c.is_conditional)
    logger.summary(
        total=len(checks),
        mandatory=mandatory_count,
        conditional=conditional_count,
        generated=generated_total,
        warnings=placeholder_count,
        errors=0
    )

    print(f"\n=== 生成完成（v2 + v0.7参数展开）===")
    print(f"总检查项: {len(checks)}")
    print(f"生成SQL总数: {len(sql_list)} 份")
    print(f"  使用SQL模板: {template_count}")
    print(f"  拼装生成: {assembled_count}")
    print(f"  参数展开新增: {param_expanded_count} 份")
    print(f"  参数未展开(仅个性化): {param_no_expand_count}")
    print(f"  需人工补充: {placeholder_count}")
    print(f"  覆盖率: {(len(sql_list) - placeholder_count) / len(sql_list) * 100:.1f}%")
    print(f"输出目录: {output_dir}")
    print(f"日志文件: {log_path}")


if __name__ == "__main__":
    main()